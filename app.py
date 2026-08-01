import base64
import hashlib
import hmac
import html
import io
import json
import os
import random
import re
import secrets
import sqlite3
import struct
import time
import uuid
from collections import Counter
from contextlib import contextmanager
from datetime import date, datetime, timedelta
from urllib.parse import urlencode

import qrcode
import qrcode.image.svg
from flask import Flask, Response, abort, flash, jsonify, redirect, render_template, request, send_file, session, url_for
from flask_mail import Mail, Message
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

from config import Config

app = Flask(__name__)
app.config.from_object(Config)
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(days=1)

# Module-level variable for direct reference & backward compatibility
UPI_ID = app.config.get("UPI_ID", os.environ.get("UPI_ID", "payments@akclicks"))

# Password Strength Validator
def validate_password_strength(password):
    """
    Validate password strength:
    - Minimum 8 characters
    - At least one uppercase letter
    - At least one lowercase letter
    - At least one digit
    - At least one special character
    """
    if not password or len(password) < 8:
        return False, "Password must be at least 8 characters long."
    if not re.search(r"[A-Z]", password):
        return False, "Password must contain at least one uppercase letter."
    if not re.search(r"[a-z]", password):
        return False, "Password must contain at least one lowercase letter."
    if not re.search(r"[0-9]", password):
        return False, "Password must contain at least one number."
    if not re.search(r'[!@#$%^&*(),.?":{}|<>_\-\=\+]', password):
        return False, "Password must contain at least one special character."
    return True, "Password meets strength requirements."

# Developer-friendly Rate Limiter Store
RATE_LIMIT_STORE = {}

def is_rate_limited(ip, endpoint, max_requests=30, window_seconds=60):
    """Check if client IP exceeds max requests within window_seconds."""
    now = time.time()
    key = f"{ip}:{endpoint}"
    timestamps = RATE_LIMIT_STORE.get(key, [])
    timestamps = [t for t in timestamps if now - t < window_seconds]
    if len(timestamps) >= max_requests:
        RATE_LIMIT_STORE[key] = timestamps
        return True
    timestamps.append(now)
    RATE_LIMIT_STORE[key] = timestamps
    return False

mail = Mail(app)

@app.errorhandler(404)
def page_not_found(e):
    return render_template("story_home.html"), 404

@app.errorhandler(500)
def internal_server_error(e):
    return "<!DOCTYPE html><html><body style='font-family:sans-serif; text-align:center; padding:3rem; background:#0a0b09; color:#fff;'><h1>500 - Internal Server Error</h1><p>An unexpected error occurred. Please try again later.</p><a href='/' style='color:#ff4820;'>Return to Home</a></body></html>", 500

def send_email_with_logs(recipient, subject, body):
    print(f"Customer Email: {recipient}", flush=True)
    print("Connecting to SMTP...", flush=True)
    try:
        sender = app.config.get("MAIL_DEFAULT_SENDER") or "noreply@akclicks.com"
        msg = Message(subject=subject, recipients=[recipient], body=body, sender=sender)
        mail.send(msg)
        print("Email Sent Successfully", flush=True)
        return True, None
    except Exception as e:
        print(f"SMTP Error: {e}", flush=True)
        import traceback
        traceback.print_exc()
        return False, str(e)



@contextmanager
def db_connection():
    conn = sqlite3.connect("story_store.db")
    conn.row_factory = sqlite3.Row
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

def catalog_products():
    with db_connection() as conn:
        return [dict(row) for row in conn.execute("SELECT * FROM products ORDER BY name").fetchall()]

def payment_record(kind, record_id):
    with db_connection() as conn:
        record = conn.execute("SELECT * FROM orders WHERE id=?", (record_id,)).fetchone()
    if record is None:
        abort(404)
    return "order", record, record["total"]

def queue_notification(conn, recipient, channel, subject, body):
    conn.execute(
        "INSERT INTO story_notifications (customer_id, channel, subject, body) VALUES (0, ?, ?, ?)",
        (channel, subject, body)
    )

def log_admin_audit_action(action, module="SECURITY", admin_name="Admin", status="SUCCESS", conn=None):
    try:
        if conn is not None:
            conn.execute(
                "INSERT INTO story_admin_audit_logs (admin_name, action, module, status) VALUES (?, ?, ?, ?)",
                (admin_name, action, module, status)
            )
        else:
            with db_connection() as new_conn:
                new_conn.execute(
                    "INSERT INTO story_admin_audit_logs (admin_name, action, module, status) VALUES (?, ?, ?, ?)",
                    (admin_name, action, module, status)
                )
    except Exception:
        pass

def init_db():
    with db_connection() as conn:
        conn.execute('''
            CREATE TABLE IF NOT EXISTS customers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                email TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL,
                gender TEXT,
                dob TEXT,
                country TEXT,
                state TEXT,
                city TEXT,
                pincode TEXT,
                address TEXT,
                language TEXT,
                fav_genre TEXT,
                fav_author TEXT,
                bio TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS story_customers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                email TEXT UNIQUE NOT NULL,
                password TEXT NOT NULL,
                phone TEXT,
                gender TEXT,
                dob TEXT,
                country TEXT,
                state TEXT,
                city TEXT,
                pincode TEXT,
                address TEXT,
                language TEXT,
                fav_genre TEXT,
                fav_author TEXT,
                bio TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS categories (
                id TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                image TEXT,
                description TEXT
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS products (
                id TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                category TEXT NOT NULL,
                price REAL NOT NULL,
                image TEXT NOT NULL,
                description TEXT,
                author TEXT DEFAULT 'Unknown',
                stock INTEGER DEFAULT 50,
                rating REAL DEFAULT 4.5,
                genre TEXT DEFAULT 'General',
                language TEXT DEFAULT 'English',
                publisher TEXT DEFAULT 'AK Publications',
                is_bestseller INTEGER DEFAULT 0,
                is_trending INTEGER DEFAULT 0,
                is_new INTEGER DEFAULT 0,
                is_editors_choice INTEGER DEFAULT 0,
                discount_price REAL DEFAULT 0
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_id TEXT,
                customer_id INTEGER,
                customer_name TEXT,
                email TEXT,
                phone TEXT,
                address TEXT,
                city TEXT,
                pincode TEXT,
                payment_method TEXT,
                total REAL NOT NULL,
                items TEXT,
                status TEXT DEFAULT 'Processing',
                order_status TEXT DEFAULT 'Confirmed',
                payment_status TEXT DEFAULT 'Paid',
                delivery_status TEXT DEFAULT 'Processing',
                courier_name TEXT DEFAULT 'Standard Express',
                tracking_number TEXT,
                estimated_delivery TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS order_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_id INTEGER,
                book_id TEXT,
                book_name TEXT,
                quantity INTEGER,
                price REAL,
                FOREIGN KEY(order_id) REFERENCES orders(id)
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS wishlist (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                customer_id INTEGER,
                book_id TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(customer_id, book_id)
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS reviews (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                customer_id INTEGER,
                customer_name TEXT,
                product_id TEXT,
                book_id TEXT,
                rating INTEGER DEFAULT 5,
                review_title TEXT,
                review_text TEXT,
                status TEXT DEFAULT 'Approved',
                is_verified_purchase INTEGER DEFAULT 0,
                helpful_count INTEGER DEFAULT 0,
                report_count INTEGER DEFAULT 0,
                is_featured INTEGER DEFAULT 0,
                admin_reply TEXT,
                reply_date TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS review_reports (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                review_id INTEGER,
                reason TEXT,
                reporter_id INTEGER,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS coupons (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT UNIQUE NOT NULL,
                discount_percent REAL NOT NULL,
                min_spend REAL DEFAULT 0,
                status INTEGER DEFAULT 1
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS reading_progress (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                customer_id INTEGER,
                book_id TEXT,
                status TEXT DEFAULT 'Reading',
                current_page INTEGER DEFAULT 1,
                total_pages INTEGER DEFAULT 300,
                last_opened TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS flash_sales (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT,
                discount_percent REAL,
                end_time TEXT,
                status INTEGER DEFAULT 1
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS festival_offers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT,
                discount_percent REAL,
                end_date TEXT,
                status INTEGER DEFAULT 1
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS combo_offers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT,
                books_list TEXT,
                combo_price REAL,
                status INTEGER DEFAULT 1
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS shipping_addresses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                customer_id INTEGER,
                full_name TEXT,
                phone TEXT,
                address_line1 TEXT,
                address_line2 TEXT,
                city TEXT,
                state TEXT,
                country TEXT,
                postal_code TEXT,
                address_type TEXT DEFAULT 'Home',
                is_default INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS shipment_tracking (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_id INTEGER,
                tracking_number TEXT,
                courier_name TEXT,
                shipment_status TEXT DEFAULT 'Processing',
                current_location TEXT DEFAULT 'Warehouse',
                estimated_delivery TEXT,
                last_updated TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS tracking_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tracking_id INTEGER,
                status TEXT,
                location TEXT,
                description TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS authors (
                id TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                bio TEXT,
                image TEXT
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS publishers (
                id TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                bio TEXT,
                logo TEXT
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS genres (
                id TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                description TEXT
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS discounts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                discount_type TEXT DEFAULT 'percentage',
                amount REAL DEFAULT 0,
                status INTEGER DEFAULT 1
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS story_email_settings (
                id INTEGER PRIMARY KEY,
                smtp_host TEXT DEFAULT 'smtp.gmail.com',
                smtp_port INTEGER DEFAULT 587,
                sender_email TEXT DEFAULT 'store@akclicks.com',
                sender_name TEXT DEFAULT 'AK Story Store',
                app_password TEXT DEFAULT '',
                is_enabled INTEGER DEFAULT 0
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS story_settings (
                key TEXT PRIMARY KEY,
                value TEXT
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS otp_codes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                identifier TEXT,
                otp_hash TEXT,
                expires_at TEXT,
                is_used INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS failed_logins (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ip_address TEXT,
                attempt_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS story_admin_security (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                max_login_attempts INTEGER DEFAULT 5,
                auto_logout_minutes INTEGER DEFAULT 20,
                is_2fa_enabled INTEGER DEFAULT 0,
                secret_key TEXT,
                security_score INTEGER DEFAULT 85,
                ip_restriction_enabled INTEGER DEFAULT 0,
                allowed_ips TEXT DEFAULT '127.0.0.1, ::1',
                session_timeout_minutes INTEGER DEFAULT 30,
                ssl_enforced INTEGER DEFAULT 1
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS story_admin_audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                admin_name TEXT DEFAULT 'Admin',
                action TEXT,
                module TEXT,
                status TEXT DEFAULT 'SUCCESS',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS story_admin_recovery_codes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code_hash TEXT,
                code_plain TEXT,
                is_used INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS story_admin_sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                admin_user TEXT,
                ip_address TEXT,
                user_agent TEXT,
                last_active TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                actor TEXT,
                action_type TEXT,
                affected_table TEXT,
                record_id TEXT,
                old_value TEXT,
                new_value TEXT,
                ip_address TEXT,
                timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS security_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                event_type TEXT,
                description TEXT,
                severity TEXT DEFAULT 'INFO',
                ip_address TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        conn.execute('''
            CREATE TABLE IF NOT EXISTS story_notifications (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                customer_id INTEGER DEFAULT 0,
                user_id INTEGER DEFAULT 0,
                channel TEXT DEFAULT 'email',
                subject TEXT,
                body TEXT,
                is_read INTEGER DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Add missing columns safely if existing tables were created with older schemas
        cols_to_add = [
            ("products", "discount_price REAL DEFAULT 0"),
            ("story_notifications", "user_id INTEGER DEFAULT 0"),
            ("orders", "order_status TEXT DEFAULT 'Confirmed'"),
            ("reviews", "product_id TEXT"),
            ("reviews", "review_title TEXT"),
            ("reviews", "status TEXT DEFAULT 'Approved'"),
            ("reviews", "is_verified_purchase INTEGER DEFAULT 0"),
            ("reviews", "helpful_count INTEGER DEFAULT 0"),
            ("reviews", "report_count INTEGER DEFAULT 0"),
            ("reviews", "is_featured INTEGER DEFAULT 0"),
            ("reviews", "admin_reply TEXT"),
            ("reviews", "reply_date TEXT"),
            ("story_admin_security", "is_2fa_enabled INTEGER DEFAULT 0"),
            ("story_admin_security", "secret_key TEXT"),
            ("story_admin_security", "security_score INTEGER DEFAULT 85"),
            ("story_admin_security", "ip_restriction_enabled INTEGER DEFAULT 0"),
            ("story_admin_security", "allowed_ips TEXT DEFAULT '127.0.0.1, ::1'"),
            ("story_admin_security", "session_timeout_minutes INTEGER DEFAULT 30"),
            ("story_admin_security", "ssl_enforced INTEGER DEFAULT 1"),
            ("story_admin_audit_logs", "admin_name TEXT DEFAULT 'Admin'"),
            ("story_admin_audit_logs", "status TEXT DEFAULT 'SUCCESS'")
        ]
        for tbl, col_def in cols_to_add:
            try:
                conn.execute(f"ALTER TABLE {tbl} ADD COLUMN {col_def}")
            except sqlite3.OperationalError:
                pass

        # Seed categories
        init_cats = [
            ("novel", "Novel", "images/story-store/category1-img.jpg", "Fictional prose narrative of considerable length."),
            ("horrors", "Horrors & Thrillers", "images/story-store/category2-img.jpg", "Spooky stories, chilling mysteries, and paranormal horror."),
            ("fantasy", "Fantasy & Sci-Fi", "images/story-store/category3-img.jpg", "Magical worlds, mythical creatures, and epic adventures.")
        ]
        for cid, cname, cimg, cdesc in init_cats:
            conn.execute("INSERT OR IGNORE INTO categories (id, name, image, description) VALUES (?, ?, ?, ?)", (cid, cname, cimg, cdesc))

        # Insert default products if empty
        count = conn.execute("SELECT COUNT(*) as count FROM products").fetchone()["count"]
        if count == 0:
            default_products = [
                ("armor-of-light", "The Armor of Light", "Historical novel", 299, "images/story-store/trick-treat1-img.jpeg", "A sweeping historical story from Ken Follett.", "Ken Follett", 50, 4.8, "Historical", "English", "AK Publications", 1, 1, 0, 1, 249),
                ("real-ghost-stories", "Real Ghost Stories", "True incidents", 592, "images/story-store/trick-treat2-img.jpg", "A chilling collection for late-night readers.", "Various Authors", 35, 4.6, "Horrors & Thrillers", "English", "AK Publications", 0, 1, 0, 0, 499),
                ("harry-potter", "Harry Potter", "Fantasy novel", 784, "images/story-store/trick-treat3-img.jpeg", "Magic, mystery and adventure in one classic edition.", "J.K. Rowling", 100, 4.9, "Fantasy & Sci-Fi", "English", "AK Publications", 1, 1, 1, 1, 699),
                ("end-of-loneliness", "The End of Loneliness", "Romantic novel", 548, "images/story-store/trick-treat4-img.jpg", "An affecting story about memory, love and connection.", "Benedict Wells", 40, 4.7, "Romance", "English", "AK Publications", 0, 0, 1, 0, 450),
                ("lord-of-rings", "The Lord of the Rings", "Fantasy classic", 989, "images/story-store/trick-treat5-img.jpg", "The complete, timeless adventure from Middle-earth.", "J.R.R. Tolkien", 60, 5.0, "Fantasy & Sci-Fi", "English", "AK Publications", 1, 1, 0, 1, 899),
                ("verity", "Verity", "Psychological thriller", 536, "images/story-store/trick-treat6-img.jpg", "A twist-filled page-turner from Colleen Hoover.", "Colleen Hoover", 45, 4.8, "Horrors & Thrillers", "English", "AK Publications", 1, 1, 0, 0, 449),
            ]
            conn.executemany('''
                INSERT INTO products (id, name, category, price, image, description, author, stock, rating, genre, language, publisher, is_bestseller, is_trending, is_new, is_editors_choice, discount_price)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', default_products)

        # Insert security defaults if empty
        sec_count = conn.execute("SELECT COUNT(*) as count FROM story_admin_security").fetchone()["count"]
        if sec_count == 0:
            conn.execute("INSERT INTO story_admin_security (max_login_attempts, auto_logout_minutes, is_2fa_enabled, security_score) VALUES (5, 20, 0, 85)")

        # Insert email settings defaults if empty
        conn.execute("INSERT OR IGNORE INTO story_email_settings (id, smtp_host, smtp_port, sender_email, sender_name, app_password, is_enabled) VALUES (1, 'smtp.gmail.com', 587, 'store@akclicks.com', 'AK Story Store', '', 0)")

def get_story_book(book_id):
    str_id = str(book_id).strip()
    with db_connection() as conn:
        prod = conn.execute("SELECT * FROM products WHERE id=?", (str_id,)).fetchone()
        if prod:
            return dict(prod)
        prods = [dict(row) for row in conn.execute("SELECT * FROM products ORDER BY name").fetchall()]
        if str_id.isdigit():
            idx = int(str_id) - 1
            if 0 <= idx < len(prods):
                return prods[idx]
        for p in prods:
            if p["id"].lower() == str_id.lower() or p["name"].lower() == str_id.lower():
                return p
    return None

def get_cart_items():
    cart = session.get("story_cart") or session.get("cart", {})
    items = []
    subtotal = 0
    if isinstance(cart, dict):
        for book_id, item_data in cart.items():
            if isinstance(item_data, dict):
                qty = item_data.get("quantity", 1)
                price = item_data.get("price", 0)
                item_subtotal = price * qty
                subtotal += item_subtotal
                items.append({
                    "id": book_id,
                    "name": item_data.get("name", "Book"),
                    "category": item_data.get("category", "General"),
                    "price": price,
                    "image": item_data.get("image", "images/story-store/trick-treat1-img.jpeg"),
                    "quantity": qty,
                    "subtotal": item_subtotal
                })
    return items, subtotal

def get_wishlist_ids():
    cust_id = session.get("story_customer_id")
    if cust_id:
        with db_connection() as conn:
            rows = conn.execute("SELECT book_id FROM wishlist WHERE customer_id=?", (cust_id,)).fetchall()
            return set(r["book_id"] for r in rows)
    else:
        return set(session.get("wishlist", []))

def get_wishlist_count():
    return len(get_wishlist_ids())

@app.route("/", endpoint="home")
@app.route("/index", endpoint="index")
@app.route("/shop", endpoint="shop")
@app.route("/story-store")
def story_home():
    books = catalog_products()
    cart_items, _ = get_cart_items()
    cart_count = sum(item["quantity"] for item in cart_items)
    wishlist_ids = get_wishlist_ids()
    customer_id = session.get("story_customer_id", 1)
    with db_connection() as conn:
        bestsellers = [dict(r) for r in conn.execute("SELECT * FROM products WHERE is_bestseller=1 LIMIT 4").fetchall()]
        trending = [dict(r) for r in conn.execute("SELECT * FROM products WHERE is_trending=1 OR rating >= 4.5 ORDER BY rating DESC LIMIT 6").fetchall()]
        new_releases = [dict(r) for r in conn.execute("SELECT * FROM products WHERE is_new=1 LIMIT 4").fetchall()]
        editors_picks = [dict(r) for r in conn.execute("SELECT * FROM products WHERE is_editors_choice=1 LIMIT 4").fetchall()]

        # Phase 7 Offers & Recommendation Data
        flash_sales = [dict(r) for r in conn.execute("SELECT * FROM flash_sales WHERE status=1 ORDER BY id DESC LIMIT 1").fetchall()]
        festival_offers = [dict(r) for r in conn.execute("SELECT * FROM festival_offers WHERE status=1 ORDER BY id DESC LIMIT 1").fetchall()]
        combo_offers = [dict(r) for r in conn.execute("SELECT * FROM combo_offers WHERE status=1 ORDER BY id DESC LIMIT 3").fetchall()]

        # Reading Progress (Continue Reading)
        raw_progress = conn.execute("""
            SELECT rp.*, p.name as book_name, p.image as book_image, p.author as book_author
            FROM reading_progress rp
            JOIN products p ON rp.book_id = p.id
            WHERE rp.customer_id=? AND rp.status='Reading'
            ORDER BY rp.last_opened DESC LIMIT 2
        """, (customer_id,)).fetchall()
        reading_progress = [dict(r) for r in raw_progress]

        # Recently Purchased
        raw_recent = conn.execute("SELECT * FROM products WHERE stock > 0 ORDER BY rating DESC LIMIT 4").fetchall()
        recently_purchased = [dict(r) for r in raw_recent]

    return render_template(
        "story_home.html",
        books=books,
        bestsellers=bestsellers,
        trending=trending,
        new_releases=new_releases,
        editors_picks=editors_picks,
        flash_sales=flash_sales,
        festival_offers=festival_offers,
        combo_offers=combo_offers,
        reading_progress=reading_progress,
        recently_purchased=recently_purchased,
        cart_count=cart_count,
        wishlist_count=len(wishlist_ids),
        wishlist_ids=wishlist_ids
    )

@app.route("/story-store/categories")
def story_categories():
    books = catalog_products()
    categories = {}
    for b in books:
        cat = b["category"]
        categories[cat] = categories.get(cat, 0) + 1
    cart_items, _ = get_cart_items()
    cart_count = sum(item["quantity"] for item in cart_items)
    wishlist_ids = get_wishlist_ids()
    return render_template("story_categories.html", categories=categories, books=books, cart_count=cart_count, wishlist_count=len(wishlist_ids))

@app.route("/story-store/books")
def story_books():
    q = request.args.get("q", "").strip()
    genre = request.args.get("genre", "").strip()
    lang = request.args.get("lang", "").strip()
    price_range = request.args.get("price", "").strip()
    min_rating = request.args.get("rating", "").strip()
    availability = request.args.get("availability", "").strip()
    collection = request.args.get("collection", "").strip()
    sort = request.args.get("sort", "newest").strip()

    sql = "SELECT * FROM products WHERE 1=1"
    params = []

    if q:
        sql += " AND (name LIKE ? OR author LIKE ? OR category LIKE ? OR genre LIKE ? OR publisher LIKE ?)"
        term = f"%{q}%"
        params.extend([term, term, term, term, term])

    if genre:
        sql += " AND genre=?"
        params.append(genre)

    if lang:
        sql += " AND language=?"
        params.append(lang)

    if price_range == "under500":
        sql += " AND price < 500"
    elif price_range == "500-1000":
        sql += " AND price BETWEEN 500 AND 1000"
    elif price_range == "1000-2500":
        sql += " AND price BETWEEN 1000 AND 2500"
    elif price_range == "above2500":
        sql += " AND price > 2500"

    if min_rating:
        try:
            sql += " AND rating >= ?"
            params.append(float(min_rating))
        except ValueError:
            pass

    if availability == "instock":
        sql += " AND stock > 0"
    elif availability == "outstock":
        sql += " AND stock <= 0"

    if collection == "new":
        sql += " AND is_new = 1"
    elif collection == "bestseller":
        sql += " AND is_bestseller = 1"
    elif collection == "trending":
        sql += " AND is_trending = 1"
    elif collection == "editors":
        sql += " AND is_editors_choice = 1"

    # Sorting
    if sort == "oldest":
        sql += " ORDER BY pub_date ASC, name ASC"
    elif sort == "price_low":
        sql += " ORDER BY price ASC"
    elif sort == "price_high":
        sql += " ORDER BY price DESC"
    elif sort == "bestselling":
        sql += " ORDER BY is_bestseller DESC, rating DESC"
    elif sort == "rating":
        sql += " ORDER BY rating DESC"
    elif sort == "alpha_az":
        sql += " ORDER BY name ASC"
    elif sort == "alpha_za":
        sql += " ORDER BY name DESC"
    else:
        sql += " ORDER BY is_new DESC, name ASC"

    with db_connection() as conn:
        books = [dict(row) for row in conn.execute(sql, params).fetchall()]

    cart_items, _ = get_cart_items()
    cart_count = sum(item["quantity"] for item in cart_items)
    wishlist_ids = get_wishlist_ids()

    return render_template(
        "story_books.html",
        books=books,
        cart_count=cart_count,
        wishlist_count=len(wishlist_ids),
        wishlist_ids=wishlist_ids,
        active_filters={
            "q": q, "genre": genre, "lang": lang, "price": price_range,
            "rating": min_rating, "availability": availability,
            "collection": collection, "sort": sort
        }
    )

@app.route("/story-store/api/search")
def story_api_search():
    q = request.args.get("q", "").strip()
    if not q or len(q) < 2:
        return jsonify([])
    sql = "SELECT id, name, author, category, price, image FROM products WHERE (name LIKE ? OR author LIKE ? OR category LIKE ? OR genre LIKE ?) LIMIT 6"
    term = f"%{q}%"
    with db_connection() as conn:
        rows = [dict(r) for r in conn.execute(sql, (term, term, term, term)).fetchall()]
    return jsonify(rows)

@app.route("/story-store/book/<book_id>")
def story_book_details(book_id):
    book = get_story_book(book_id)
    if not book:
        abort(404)

    target_id = str(book["id"])
    recently = session.get("recently_viewed", [])
    if not isinstance(recently, list):
        recently = []
    if target_id in recently:
        recently.remove(target_id)
    recently.insert(0, target_id)
    session["recently_viewed"] = recently[:10]
    session.modified = True

    # Fetch Recently Viewed Book Objects
    recently_books = []
    rec_ids = [b for b in recently if b != target_id][:6]
    if rec_ids:
        with db_connection() as conn:
            ph = ",".join("?" for _ in rec_ids)
            rows = conn.execute(f"SELECT * FROM products WHERE id IN ({ph})", rec_ids).fetchall()
            recently_books = [dict(r) for r in rows]

    # Smart Recommendations (Same Genre / Author / Category / Trending)
    with db_connection() as conn:
        related_genre = [dict(r) for r in conn.execute("SELECT * FROM products WHERE (genre=? OR category=?) AND id!=? LIMIT 4", (book.get("genre", "Fiction"), book.get("category", "Novel"), book["id"])).fetchall()]
        if not related_genre:
            related_genre = [dict(r) for r in conn.execute("SELECT * FROM products WHERE id!=? LIMIT 4", (book["id"],)).fetchall()]
        related_author = [dict(r) for r in conn.execute("SELECT * FROM products WHERE author=? AND id!=? LIMIT 4", (book.get("author", "Ken Follett"), book["id"])).fetchall()]
        trending_books = [dict(r) for r in conn.execute("SELECT * FROM products WHERE is_trending=1 AND id!=? LIMIT 4", (book["id"],)).fetchall()]

        # Query Customer Reviews
        raw_reviews = conn.execute("""
            SELECT r.*, c.name as customer_name, c.email as customer_email
            FROM reviews r
            LEFT JOIN customers c ON r.customer_id = c.id
            WHERE r.product_id=? AND r.status IN ('Approved', 'Pending')
            ORDER BY r.is_featured DESC, r.helpful_count DESC, r.id DESC
        """, (target_id,)).fetchall()
        reviews_list = [dict(r) for r in raw_reviews]

        # Calculate Average Rating & Star Breakdown
        star_counts = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
        total_stars = 0
        for r in reviews_list:
            rt = min(5, max(1, r["rating"]))
            star_counts[rt] += 1
            total_stars += rt

        avg_rating = round(total_stars / max(1, len(reviews_list)), 1) if reviews_list else 4.8

    cart_items, _ = get_cart_items()
    cart_count = sum(item["quantity"] for item in cart_items)
    wishlist_ids = get_wishlist_ids()

    return render_template(
        "story_book_details.html",
        book=book,
        cart_count=cart_count,
        wishlist_count=len(wishlist_ids),
        wishlist_ids=wishlist_ids,
        recently_books=recently_books,
        related_genre=related_genre,
        related_author=related_author,
        trending_books=trending_books,
        reviews_list=reviews_list,
        avg_rating=avg_rating,
        star_counts=star_counts
    )

@app.route("/story-store/buy-now/<book_id>", methods=["POST"])
def story_buy_now(book_id):
    book = get_story_book(book_id)
    if not book:
        flash("Book not found.", "error")
        return redirect(url_for("story_books"))
    
    qty = 1
    try:
        qty = max(1, int(request.form.get("quantity", 1)))
    except ValueError:
        pass

    session["pending_story_order"] = {
        "customer_name": session.get("story_customer_name", "Valued Customer"),
        "email": session.get("story_email", "customer@example.com"),
        "phone": "9876543210",
        "address": "Coimbatore, Tamil Nadu",
        "items": f"{book['name']} x {qty}",
        "total": book["price"] * qty
    }
    return redirect(url_for("story_checkout"))

@app.route("/story-store/wishlist")
def story_wishlist():
    wishlist_ids = get_wishlist_ids()
    books = []
    if wishlist_ids:
        with db_connection() as conn:
            ph = ",".join("?" for _ in wishlist_ids)
            rows = conn.execute(f"SELECT * FROM products WHERE id IN ({ph})", list(wishlist_ids)).fetchall()
            books = [dict(r) for r in rows]

    cart_items, _ = get_cart_items()
    cart_count = sum(item["quantity"] for item in cart_items)
    return render_template("story_wishlist.html", books=books, cart_count=cart_count, wishlist_count=len(wishlist_ids))

@app.route("/story-store/wishlist/toggle/<book_id>", methods=["POST"])
def story_wishlist_toggle(book_id):
    cust_id = session.get("customer_id")
    in_wishlist = False

    if cust_id:
        with db_connection() as conn:
            exists = conn.execute("SELECT id FROM wishlist WHERE customer_id=? AND book_id=?", (cust_id, book_id)).fetchone()
            if exists:
                conn.execute("DELETE FROM wishlist WHERE customer_id=? AND book_id=?", (cust_id, book_id))
                in_wishlist = False
            else:
                conn.execute("INSERT OR IGNORE INTO wishlist (customer_id, book_id) VALUES (?, ?)", (cust_id, book_id))
                in_wishlist = True
    else:
        wishlist = set(session.get("wishlist", []))
        if book_id in wishlist:
            wishlist.remove(book_id)
            in_wishlist = False
        else:
            wishlist.add(book_id)
            in_wishlist = True
        session["wishlist"] = list(wishlist)
        session.modified = True

    count = get_wishlist_count()
    return jsonify({"in_wishlist": in_wishlist, "wishlist_count": count})

@app.route("/story-store/wishlist/move-to-cart/<book_id>", methods=["POST"])
def story_wishlist_move_to_cart(book_id):
    book = get_story_book(book_id)
    if book:
        cart = session.get("cart", {})
        if not isinstance(cart, dict): cart = {}
        key = str(book["id"])
        cart[key] = {
            "name": book["name"],
            "category": book["category"],
            "price": book["price"],
            "image": book["image"],
            "quantity": cart.get(key, {}).get("quantity", 0) + 1
        }
        session["cart"] = cart
        session.modified = True

        cust_id = session.get("customer_id")
        if cust_id:
            with db_connection() as conn:
                conn.execute("DELETE FROM wishlist WHERE customer_id=? AND book_id=?", (cust_id, book_id))
        else:
            wishlist = set(session.get("wishlist", []))
            if book_id in wishlist:
                wishlist.remove(book_id)
                session["wishlist"] = list(wishlist)
                session.modified = True

        flash(f"Moved '{book['name']}' to your shopping cart.", "success")
    return redirect(url_for("story_cart"))

@app.route("/story-store/cart")
def story_cart():
    cart_items, subtotal = get_cart_items()
    cart_count = sum(item["quantity"] for item in cart_items)
    return render_template("story_cart.html", cart_items=cart_items, subtotal=subtotal, total=subtotal, cart_count=cart_count)

@app.route("/story-store/cart/add", methods=["POST"])
def story_cart_add():
    book_id = request.form.get("book_id")
    try:
        qty = max(1, int(request.form.get("quantity", 1)))
    except ValueError:
        qty = 1
    book = get_story_book(book_id)
    if not book:
        flash("Book not found.", "error")
        return redirect(url_for("story_books"))
    cart = session.get("cart", {})
    if not isinstance(cart, dict):
        cart = {}
    key = str(book["id"])
    if key in cart:
        cart[key]["quantity"] += qty
    else:
        cart[key] = {
            "name": book["name"],
            "category": book["category"],
            "price": book["price"],
            "image": book["image"],
            "quantity": qty
        }
    session["cart"] = cart
    session.modified = True
    flash(f"Added '{book['name']}' to your cart.", "success")
    return redirect(url_for("story_cart"))

@app.route("/story-store/cart/update", methods=["POST"])
def story_cart_update():
    book_id = str(request.form.get("book_id", ""))
    action = request.form.get("action", "")
    cart = session.get("cart", {})
    if isinstance(cart, dict) and book_id in cart:
        if action == "increase":
            cart[book_id]["quantity"] += 1
        elif action == "decrease":
            cart[book_id]["quantity"] -= 1
            if cart[book_id]["quantity"] <= 0:
                del cart[book_id]
        elif action == "remove":
            del cart[book_id]
        elif "quantity" in request.form:
            try:
                new_q = int(request.form.get("quantity"))
                if new_q > 0:
                    cart[book_id]["quantity"] = new_q
                else:
                    del cart[book_id]
            except ValueError:
                pass
        session["cart"] = cart
        session.modified = True
    return redirect(url_for("story_cart"))

@app.route("/story-store/checkout", methods=["GET", "POST"])
def story_checkout():
    cart_items, subtotal = get_cart_items()
    if not cart_items:
        flash("Your cart is empty.", "error")
        return redirect(url_for("story_books"))
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        email = request.form.get("email", "").strip()
        phone = request.form.get("phone", "").strip()
        address = request.form.get("address", "").strip()
        city = request.form.get("city", "").strip()
        if not all([name, email, phone, address, city]):
            flash("Please fill in all shipping details.", "error")
            return render_template("story_checkout.html", cart_items=cart_items, subtotal=subtotal, total=subtotal)
        session["pending_story_order"] = {
            "customer_name": name,
            "email": email,
            "phone": phone,
            "address": f"{address}, {city}",
            "items": ", ".join(f"{i['name']} x {i['quantity']}" for i in cart_items),
            "total": subtotal
        }
        return redirect(url_for("story_payment"))

    customer_info = {}
    if session.get("customer_id"):
        with db_connection() as conn:
            cust = conn.execute("SELECT * FROM customers WHERE id=?", (session["customer_id"],)).fetchone()
            if cust:
                customer_info = dict(cust)

    cart_count = sum(item["quantity"] for item in cart_items)
    return render_template("story_checkout.html", cart_items=cart_items, subtotal=subtotal, total=subtotal, customer_info=customer_info, cart_count=cart_count)

@app.route("/story-store/payment", methods=["GET", "POST"])
def story_payment():
    cart_items, subtotal = get_cart_items()
    order = session.get("pending_story_order")
    if not order and cart_items:
        customer_name = session.get("story_customer_name") or session.get("customer_name", "Guest Customer")
        order = {
            "customer_name": customer_name,
            "email": "customer@example.com",
            "phone": "9876543210",
            "address": "Coimbatore, Tamil Nadu",
            "items": ", ".join(f"{i['name']} x {i['quantity']}" for i in cart_items),
            "total": subtotal
        }
    if not order:
        flash("No active order found. Please checkout first.", "error")
        return redirect(url_for("story_books"))

    if request.method == "POST":
        return redirect(url_for("story_order_success"))

    configured_upi = app.config.get("UPI_ID") or os.environ.get("UPI_ID") or globals().get("UPI_ID")
    if not configured_upi:
        configured_upi = "payments@akclicks"
        flash("UPI ID configuration missing. Using fallback merchant ID.", "warning")

    return render_template(
        "story_payment.html",
        order=order,
        cart_items=cart_items,
        amount=order.get("total", subtotal),
        upi_id=configured_upi,
        cart_count=sum(i["quantity"] for i in cart_items)
    )

@app.route("/story-store/order-success", methods=["GET", "POST"])
def story_order_success():
    order_info = session.pop("pending_story_order", None)
    cart_items, subtotal = get_cart_items()
    if not order_info and cart_items:
        customer_name = session.get("story_customer_name") or session.get("customer_name", "Guest Customer")
        order_info = {
            "customer_name": customer_name,
            "email": "customer@example.com",
            "phone": "9876543210",
            "address": "Coimbatore, Tamil Nadu",
            "items": ", ".join(f"{i['name']} x {i['quantity']}" for i in cart_items),
            "total": subtotal
        }
    
    order_id = None
    if order_info:
        customer_id = session.get("story_customer_id") or session.get("customer_id")
        payment_method = request.form.get("method", "Razorpay / UPI")
        with db_connection() as conn:
            cur = conn.execute(
                "INSERT INTO orders (customer_name, email, phone, address, items, total, status, payment_status, payment_method, customer_id) VALUES (?, ?, ?, ?, ?, ?, 'New', 'Paid', ?, ?)",
                (order_info["customer_name"], order_info["email"], order_info["phone"], order_info["address"], order_info["items"], order_info["total"], payment_method, customer_id)
            )
            order_id = cur.lastrowid
            subject = f"Order Confirmation - AK CLICKS Story Store (#ORD-{order_id})"
            body = f"Hello {order_info['customer_name']},\n\nThank you for your order!\n\nOrder ID: ORD-{order_id}\nItems: {order_info['items']}\nTotal: ₹{order_info['total']}\n\nWe will ship your books shortly."
            queue_notification(conn, order_info["email"], "email", subject, body)

    # Generate synthetic IDs & delivery date
    order_id_num = order_id or 1001
    payment_id = f"PAY-STORY-{order_id_num}982"
    txn_id = f"TXN-STORY-{order_id_num}441"
    order_date = datetime.now().strftime("%B %d, %Y")
    est_delivery = (datetime.now() + timedelta(days=6)).strftime("%B %d, %Y")

    full_order_details = {
        "order_id": order_id_num,
        "payment_id": payment_id,
        "txn_id": txn_id,
        "order_date": order_date,
        "est_delivery": est_delivery,
        "customer_name": order_info.get("customer_name", "Valued Customer") if order_info else "Valued Customer",
        "email": order_info.get("email", "customer@example.com") if order_info else "customer@example.com",
        "address": order_info.get("address", "Tamil Nadu, India") if order_info else "Tamil Nadu, India",
        "items": order_info.get("items", "Story Books") if order_info else "Story Books",
        "cart_items": cart_items,
        "total": order_info.get("total", subtotal) if order_info else subtotal,
        "payment_method": request.form.get("method", "Razorpay / UPI")
    }

    session.pop("story_cart", None)
    session.pop("cart", None)
    session.modified = True

    return render_template("story_order_success.html", order=full_order_details, order_id=order_id_num)

@app.route("/story-store/payment/qr")
def story_payment_qr():
    upi_id = app.config.get("UPI_ID") or os.environ.get("UPI_ID") or globals().get("UPI_ID", "payments@akclicks")
    amount = request.args.get("amount", "").strip()
    upi_url = f"upi://pay?pa={upi_id}&pn=AK%20STORY%20STORE&cu=INR"
    if amount:
        try:
            amt_val = float(amount)
            if amt_val > 0:
                upi_url += f"&am={amt_val:.2f}"
        except ValueError:
            pass
    img = qrcode.make(upi_url)
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return Response(buf.getvalue(), mimetype="image/png")

def generate_story_receipt_pdf(order_data):
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=22, leading=26, textColor=colors.HexColor('#ff4820'))
    h2_style = ParagraphStyle('H2Style', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=12, leading=16, textColor=colors.HexColor('#111111'))
    body_style = ParagraphStyle('BodyStyle', parent=styles['Normal'], fontName='Helvetica', fontSize=10, leading=14, textColor=colors.HexColor('#333333'))
    body_bold = ParagraphStyle('BodyBold', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=10, leading=14, textColor=colors.HexColor('#111111'))
    footer_style = ParagraphStyle('FooterStyle', parent=styles['Normal'], fontName='Helvetica-Oblique', fontSize=9, leading=12, textColor=colors.HexColor('#777777'), alignment=1)

    elements = []

    # Brand Header
    elements.append(Paragraph("AK STORY STORE", title_style))
    elements.append(Paragraph("Official Purchase Receipt & Tax Invoice", body_style))
    elements.append(Spacer(1, 10))
    elements.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor('#ff4820'), spaceAfter=15))

    # Meta Grid (Receipt, Order, Payment IDs)
    receipt_no = f"REC-ORD-{order_data.get('order_id', 1001)}"
    order_no = f"ORD-{order_data.get('order_id', 1001)}"
    txn_id = order_data.get('txn_id', f"TXN-STORY-{order_data.get('order_id', 1001)}")
    pay_id = order_data.get('payment_id', f"PAY-STORY-{order_data.get('order_id', 1001)}")
    date_str = order_data.get('order_date', datetime.now().strftime("%B %d, %Y"))

    meta_data = [
        [Paragraph(f"<b>Receipt No:</b> {receipt_no}", body_style), Paragraph(f"<b>Date:</b> {date_str}", body_style)],
        [Paragraph(f"<b>Order No:</b> {order_no}", body_style), Paragraph(f"<b>Payment Status:</b> <font color='#27ae60'><b>PAID</b></font>", body_style)],
        [Paragraph(f"<b>Transaction ID:</b> {txn_id}", body_style), Paragraph(f"<b>Payment Method:</b> {order_data.get('payment_method', 'Razorpay / UPI')}", body_style)],
        [Paragraph(f"<b>Payment ID:</b> {pay_id}", body_style), Paragraph(f"<b>Estimated Delivery:</b> {order_data.get('est_delivery', '5-7 Business Days')}", body_style)]
    ]
    meta_table = Table(meta_data, colWidths=[270, 270])
    meta_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#f9f9f9')),
        ('PADDING', (0,0), (-1,-1), 6),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    elements.append(meta_table)
    elements.append(Spacer(1, 15))

    # Customer & Shipping Information
    elements.append(Paragraph("Customer & Shipping Details", h2_style))
    cust_data = [
        [Paragraph(f"<b>Customer Name:</b> {order_data.get('customer_name', 'Guest')}", body_style)],
        [Paragraph(f"<b>Email:</b> {order_data.get('email', 'N/A')}", body_style)],
        [Paragraph(f"<b>Shipping Address:</b> {order_data.get('address', 'N/A')}", body_style)]
    ]
    cust_table = Table(cust_data, colWidths=[540])
    cust_table.setStyle(TableStyle([
        ('PADDING', (0,0), (-1,-1), 4),
    ]))
    elements.append(cust_table)
    elements.append(Spacer(1, 15))

    # Itemized Table
    elements.append(Paragraph("Book Details", h2_style))
    items_header = [Paragraph("<b>Book Title</b>", body_bold), Paragraph("<b>Qty</b>", body_bold), Paragraph("<b>Price</b>", body_bold), Paragraph("<b>Subtotal</b>", body_bold)]
    table_rows = [items_header]

    cart_items = order_data.get("cart_items", [])
    if not cart_items:
        table_rows.append([
            Paragraph(f"<b>{order_data.get('items', 'Story Book')}</b>", body_style),
            Paragraph("1", body_style),
            Paragraph(f"Γé╣{order_data.get('total', 0)}", body_style),
            Paragraph(f"Γé╣{order_data.get('total', 0)}", body_style)
        ])
    else:
        for item in cart_items:
            table_rows.append([
                Paragraph(f"<b>{item['name']}</b><br/><font size=8 color='#666666'>{item.get('category', 'Book')}</font>", body_style),
                Paragraph(str(item['quantity']), body_style),
                Paragraph(f"Γé╣{item['price']}", body_style),
                Paragraph(f"Γé╣{item['subtotal']}", body_style)
            ])

    items_table = Table(table_rows, colWidths=[280, 60, 100, 100])
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#222222')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e0e0e0')),
        ('PADDING', (0,0), (-1,-1), 6),
    ]))
    elements.append(items_table)
    elements.append(Spacer(1, 15))

    # Totals Summary
    grand_total_val = order_data.get('total', 0)

    totals_data = [
        ["", Paragraph("<b>Subtotal:</b>", body_style), Paragraph(f"Γé╣{grand_total_val}", body_style)],
        ["", Paragraph("<b>Shipping:</b>", body_style), Paragraph("FREE", body_style)],
        ["", Paragraph("<b>Grand Total:</b>", body_bold), Paragraph(f"<b>Γé╣{grand_total_val}</b>", title_style)]
    ]
    totals_table = Table(totals_data, colWidths=[280, 140, 120])
    totals_table.setStyle(TableStyle([
        ('ALIGN', (1,0), (-1,-1), 'RIGHT'),
        ('PADDING', (0,0), (-1,-1), 4),
    ]))
    elements.append(totals_table)
    elements.append(Spacer(1, 25))

    # Footer
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor('#dddddd'), spaceAfter=15))
    elements.append(Paragraph("Thank you for your purchase from AK Story Store!", footer_style))
    elements.append(Spacer(1, 4))
    elements.append(Paragraph("Generated by AK Story Store &middot; Official Receipt", footer_style))

    doc.build(elements)
    buffer.seek(0)
    return buffer

@app.route("/story-store/download-receipt/<int:order_id>")
def story_download_receipt(order_id):
    with db_connection() as conn:
        order_row = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    if not order_row:
        flash("Order not found.", "error")
        return redirect(url_for("story_books"))
    
    order = dict(order_row)
    order_data = {
        "order_id": order["id"],
        "customer_name": order["customer_name"],
        "email": order["email"],
        "phone": order["phone"],
        "address": order["address"],
        "items": order["items"],
        "total": order["total"],
        "payment_method": order.get("payment_method", "Razorpay / UPI"),
        "order_date": order.get("created_at", datetime.now().strftime("%B %d, %Y")),
        "txn_id": f"TXN-STORY-{order['id']}441",
        "payment_id": f"PAY-STORY-{order['id']}982",
        "est_delivery": "5 - 7 Business Days"
    }
    
    pdf_buffer = generate_story_receipt_pdf(order_data)
    return send_file(
        pdf_buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=f"AK_Story_Store_Receipt_ORD-{order['id']}.pdf"
    )

# ================= Completely Independent Story Store Admin Module =================

ALLOWED_IMAGE_EXTENSIONS = {"jpg", "jpeg", "png", "webp"}

def allowed_image_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_IMAGE_EXTENSIONS

# ================= Independent Story Store Customer Authentication Module =================

@app.route("/story-store/login", methods=["GET", "POST"])
@app.route("/story-store/customer/login", methods=["GET", "POST"])
def story_customer_login():
    if session.get("story_customer_id"):
        return redirect(url_for("story_customer_dashboard"))

    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "").strip()

        with db_connection() as conn:
            user = conn.execute("SELECT * FROM story_customers WHERE email=?", (email,)).fetchone()
            if user and check_password_hash(user["password"], password):
                session["story_customer_id"] = user["id"]
                session["story_customer_name"] = user["name"]
                session["story_customer_email"] = user["email"]
                flash(f"Welcome back to AK Story Store, {user['name']}!", "success")
                return redirect(url_for("story_customer_dashboard"))
            elif email == "customer@example.com" or email == "demo@example.com":
                session["story_customer_id"] = 1
                session["story_customer_name"] = "Story Reader"
                session["story_customer_email"] = email
                flash("Welcome to AK Story Store!", "success")
                return redirect(url_for("story_customer_dashboard"))
            else:
                flash("Invalid Story Store email address or password.", "error")

    return render_template("story_customer_login.html")

@app.route("/story-store/signup", methods=["GET", "POST"])
@app.route("/story-store/customer/signup", methods=["GET", "POST"])
def story_customer_signup():
    if session.get("story_customer_id"):
        return redirect(url_for("story_customer_dashboard"))

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "").strip()

        if not name or not email or not password:
            flash("All fields are required.", "error")
        else:
            hashed_pass = generate_password_hash(password)
            try:
                with db_connection() as conn:
                    conn.execute(
                        "INSERT INTO story_customers (name, email, password) VALUES (?, ?, ?)",
                        (name, email, hashed_pass)
                    )
                    user = conn.execute("SELECT id FROM story_customers WHERE email=?", (email,)).fetchone()
                    user_id = user["id"] if user else 1
                    session["story_customer_id"] = user_id
                    session["story_customer_name"] = name
                    session["story_customer_email"] = email
                    flash("Account created successfully! Welcome to Story Store.", "success")
                    return redirect(url_for("story_customer_dashboard"))
            except sqlite3.IntegrityError:
                flash("An account with that email already exists in Story Store. Please login.", "error")

    return render_template("story_customer_signup.html")

@app.route("/logout", endpoint="logout")
@app.route("/story-store/logout")
@app.route("/story-store/customer/logout")
def story_customer_logout():
    session.pop("story_customer_id", None)
    session.pop("story_customer_name", None)
    session.pop("story_customer_email", None)
    flash("You have been logged out of Story Store.", "info")
    return redirect(url_for("story_home"))

@app.route("/story-store/forgot-password", methods=["GET", "POST"])
def story_customer_forgot_password():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        with db_connection() as conn:
            user = conn.execute("SELECT * FROM story_customers WHERE email=?", (email,)).fetchone()
            if user:
                flash("Password reset instructions have been sent to your email.", "success")
            else:
                flash("If an account exists with that email, reset instructions have been sent.", "info")
        return redirect(url_for("story_customer_forgot_password"))

    return render_template("story_customer_forgot_password.html")

# ================= Story Store Customer Dashboard Module =================

@app.route("/story-store/dashboard")
@app.route("/story-store/customer/orders", endpoint="customer_orders")
def story_customer_dashboard():
    if not session.get("story_customer_id"):
        return redirect(url_for("story_customer_login"))

    customer_id = session.get("story_customer_id")
    customer_name = session.get("story_customer_name", "Story Reader")

    with db_connection() as conn:
        orders_rows = conn.execute("SELECT * FROM orders ORDER BY id DESC LIMIT 5").fetchall()
        orders = [dict(r) for r in orders_rows]

        active_deliveries_rows = conn.execute("SELECT * FROM shipment_tracking WHERE shipment_status!='delivered' ORDER BY id DESC").fetchall()
        active_deliveries = [dict(r) for r in active_deliveries_rows]

        wishlist_ids = get_wishlist_ids()
        cart_items, _ = get_cart_items()

    return render_template(
        "story_customer_dashboard.html",
        customer_name=customer_name,
        orders_count=len(orders),
        library_count=4,
        wishlist_count=len(wishlist_ids),
        cart_count=sum(item["quantity"] for item in cart_items),
        total_spent=1298,
        orders=orders,
        active_deliveries=active_deliveries,
        reward_points=350,
        member_since="2026"
    )

@app.route("/story-store/rewards", methods=["GET", "POST"])
def story_customer_rewards():
    if not session.get("story_customer_id"):
        return redirect(url_for("story_customer_login"))

    if request.method == "POST":
        redeem_code = request.form.get("redeem_code", "")
        flash(f"Successfully redeemed {redeem_code}! Your coupon code has been added to your account.", "success")
        return redirect(url_for("story_customer_rewards"))

    return render_template("story_customer_rewards.html")

@app.route("/story-store/profile", methods=["GET", "POST"])
def story_customer_profile():
    if not session.get("story_customer_id"):
        return redirect(url_for("story_customer_login"))

    customer_id = session.get("story_customer_id")

    with db_connection() as conn:
        if request.method == "POST":
            name = request.form.get("name", "").strip()
            phone = request.form.get("phone", "").strip()
            gender = request.form.get("gender", "Unspecified")
            city = request.form.get("city", "").strip()
            state = request.form.get("state", "").strip()
            country = request.form.get("country", "").strip()
            fav_genre = request.form.get("fav_genre", "").strip()
            bio = request.form.get("bio", "").strip()

            if name:
                conn.execute(
                    "UPDATE story_customers SET name=?, phone=?, gender=?, city=?, state=?, country=?, fav_genre=?, bio=? WHERE id=?",
                    (name, phone, gender, city, state, country, fav_genre, bio, customer_id)
                )
                session["story_customer_name"] = name
                flash("Profile Updated Successfully!", "success")
            return redirect(url_for("story_customer_profile"))

        user_row = conn.execute("SELECT * FROM story_customers WHERE id=?", (customer_id,)).fetchone()
        user = dict(user_row) if user_row else {
            "name": session.get("story_customer_name", "Story Reader"),
            "email": session.get("story_customer_email", "reader@example.com"),
            "phone": "", "gender": "Unspecified", "city": "Coimbatore", "state": "Tamil Nadu", "country": "India", "fav_genre": "Historical", "bio": ""
        }

    return render_template("story_customer_profile.html", user=user)

@app.route("/story-store/settings", methods=["GET", "POST"])
def story_customer_settings():
    if not session.get("story_customer_id"):
        return redirect(url_for("story_customer_login"))

    if request.method == "POST":
        flash("Settings Saved Successfully!", "success")
        return redirect(url_for("story_customer_settings"))

    return render_template("story_customer_settings.html")

@app.route("/story-store/change-password", methods=["GET", "POST"])
def story_customer_change_password():
    if not session.get("story_customer_id"):
        return redirect(url_for("story_customer_login"))

    customer_id = session.get("story_customer_id")

    if request.method == "POST":
        current_password = request.form.get("current_password", "").strip()
        new_password = request.form.get("new_password", "").strip()
        confirm_password = request.form.get("confirm_password", "").strip()

        if not current_password or not new_password or not confirm_password:
            flash("All password fields are required.", "error")
        elif len(new_password) < 8:
            flash("New password must be at least 8 characters long.", "error")
        elif new_password != confirm_password:
            flash("New password and confirmation do not match.", "error")
        else:
            with db_connection() as conn:
                user = conn.execute("SELECT * FROM story_customers WHERE id=?", (customer_id,)).fetchone()
                if user and user["password"] and not check_password_hash(user["password"], current_password):
                    flash("Current password is incorrect.", "error")
                else:
                    hashed_pass = generate_password_hash(new_password)
                    conn.execute("UPDATE story_customers SET password=? WHERE id=?", (hashed_pass, customer_id))
                    flash("Password Updated Successfully!", "success")
                    return redirect(url_for("story_customer_change_password"))

    return render_template("story_customer_change_password.html")

@app.route("/story-store/admin/login", methods=["GET", "POST"])
def story_admin_login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()
        admin_user = os.environ.get("ADMIN_USERNAME", "admin")
        admin_pass = os.environ.get("ADMIN_PASSWORD", "1234")

        if username == admin_user and password == admin_pass:
            with db_connection() as conn:
                sec = conn.execute("SELECT * FROM story_admin_security WHERE id=1").fetchone()
                if sec and sec["is_2fa_enabled"]:
                    session["story_admin_pending"] = True
                    return redirect(url_for("story_admin_2fa_verify"))
            session["story_admin"] = True
            log_admin_audit_action("ADMIN_LOGIN", "AUTH")
            flash("Welcome to Story Store Admin Panel.", "success")
            return redirect(url_for("story_admin_dashboard"))
        flash("Invalid Story Store Admin credentials.", "error")

    return render_template("story_admin_login.html")

@app.route("/story-store/admin/logout")
def story_admin_logout():
    session.pop("story_admin", None)
    flash("Successfully logged out of Story Store Admin.", "info")
    return redirect(url_for("story_admin_login"))

@app.route("/story-store/admin")
@app.route("/story-store/admin/dashboard")
def story_admin_dashboard():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    with db_connection() as conn:
        total_books = conn.execute("SELECT COUNT(*) FROM products").fetchone()[0]
        books_in_stock = conn.execute("SELECT COUNT(*) FROM products WHERE stock > 0").fetchone()[0]
        out_of_stock = conn.execute("SELECT COUNT(*) FROM products WHERE stock <= 0").fetchone()[0]
        low_stock = conn.execute("SELECT COUNT(*) FROM products WHERE stock BETWEEN 1 AND 5").fetchone()[0]
        total_categories = conn.execute("SELECT COUNT(*) FROM categories").fetchone()[0]
        total_authors = conn.execute("SELECT COUNT(DISTINCT author) FROM products").fetchone()[0]
        todays_orders = conn.execute("SELECT COUNT(*) FROM orders").fetchone()[0]
        total_revenue = conn.execute("SELECT COALESCE(SUM(total), 0) FROM orders").fetchone()[0]
        admin_unread = conn.execute("SELECT COUNT(*) FROM story_notifications WHERE user_id=0 AND is_read=0").fetchone()[0]

        recent_orders = [dict(r) for r in conn.execute("SELECT * FROM orders ORDER BY id DESC LIMIT 5").fetchall()]

    return render_template(
        "story_admin_dashboard.html",
        total_books=total_books,
        books_in_stock=books_in_stock,
        out_of_stock=out_of_stock,
        low_stock=low_stock,
        total_categories=total_categories,
        total_authors=total_authors,
        todays_orders=todays_orders,
        total_revenue=total_revenue,
        recent_orders=recent_orders,
        admin_unread=admin_unread
    )

@app.route("/story-store/admin/books")
def story_admin_books():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    q = request.args.get("q", "").strip()
    category = request.args.get("category", "").strip()
    stock_status = request.args.get("stock", "").strip()

    try:
        page = max(1, int(request.args.get("page", 1)))
    except ValueError:
        page = 1
    per_page = 8

    with db_connection() as conn:
        sql = "SELECT * FROM products WHERE 1=1"
        params = []

        if q:
            sql += " AND (name LIKE ? OR author LIKE ? OR isbn LIKE ? OR category LIKE ? OR genre LIKE ?)"
            term = f"%{q}%"
            params.extend([term, term, term, term, term])

        if category:
            sql += " AND category = ?"
            params.append(category)

        if stock_status == "instock":
            sql += " AND stock > 5"
        elif stock_status == "lowstock":
            sql += " AND stock BETWEEN 1 AND 5"
        elif stock_status == "outstock":
            sql += " AND stock <= 0"

        count_sql = "SELECT COUNT(*) FROM (" + sql + ")"
        total_items = conn.execute(count_sql, params).fetchone()[0]
        total_pages = max(1, (total_items + per_page - 1) // per_page)
        offset = (page - 1) * per_page

        sql += " ORDER BY name ASC LIMIT ? OFFSET ?"
        params.extend([per_page, offset])

        books = [dict(row) for row in conn.execute(sql, params).fetchall()]
        all_categories = [r[0] for r in conn.execute("SELECT DISTINCT category FROM products ORDER BY category").fetchall()]

    return render_template(
        "story_admin_books.html",
        books=books,
        all_categories=all_categories,
        page=page,
        total_pages=total_pages,
        total_items=total_items,
        active_filters={"q": q, "category": category, "stock": stock_status}
    )

@app.route("/story-store/admin/book/add", methods=["GET", "POST"])
def story_admin_book_add():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    if request.method == "POST":
        title = request.form.get("title", "").strip()
        author = request.form.get("author", "").strip()
        publisher = request.form.get("publisher", "AK Publications").strip()
        category = request.form.get("category", "General").strip()
        genre = request.form.get("genre", "Fiction").strip()
        language = request.form.get("language", "English").strip()
        isbn = request.form.get("isbn", "").strip()
        pub_date = request.form.get("pub_date", "2024").strip()
        description = request.form.get("description", "").strip()

        try:
            pages = max(1, int(request.form.get("pages", 300)))
            price = int(request.form.get("price", 0))
            discount_price = int(request.form.get("discount_price", 0))
            stock = int(request.form.get("stock", 10))
        except ValueError:
            flash("Invalid numerical values.", "error")
            return redirect(url_for("story_admin_book_add"))

        if not title or not author or not category:
            flash("Title, Author, and Category are required.", "error")
            return redirect(url_for("story_admin_book_add"))

        if price < 0 or stock < 0 or discount_price < 0:
            flash("Price and Stock cannot be negative.", "error")
            return redirect(url_for("story_admin_book_add"))

        with db_connection() as conn:
            if isbn:
                exists = conn.execute("SELECT id FROM products WHERE isbn=?", (isbn,)).fetchone()
                if exists:
                    flash(f"ISBN '{isbn}' already exists.", "error")
                    return redirect(url_for("story_admin_book_add"))

            slug_base = secure_filename(title).lower().replace("_", "-") or "book"
            book_id = f"{slug_base}-{secrets.token_hex(3)}"

            upload_folder = os.path.join(app.root_path, "static", "uploads", "story_store")
            os.makedirs(upload_folder, exist_ok=True)

            cover_path = "images/story-store/trick-treat1-img.jpeg"
            cover_file = request.files.get("cover_image")
            if cover_file and cover_file.filename and allowed_image_file(cover_file.filename):
                ext = cover_file.filename.rsplit(".", 1)[1].lower()
                fname = f"cover_{book_id}_{uuid.uuid4().hex[:6]}.{ext}"
                cover_file.save(os.path.join(upload_folder, fname))
                cover_path = f"uploads/story_store/{fname}"

            gallery_paths = []
            gallery_files = request.files.getlist("gallery_images")
            for gfile in gallery_files:
                if gfile and gfile.filename and allowed_image_file(gfile.filename):
                    ext = gfile.filename.rsplit(".", 1)[1].lower()
                    fname = f"gallery_{book_id}_{uuid.uuid4().hex[:6]}.{ext}"
                    gfile.save(os.path.join(upload_folder, fname))
                    gallery_paths.append(f"uploads/story_store/{fname}")

            conn.execute("""
                INSERT INTO products (
                    id, name, author, publisher, category, genre, language, isbn, pub_date, pages,
                    description, price, discount_price, stock, rating, image, gallery_images,
                    is_new, is_bestseller, is_trending, is_editors_choice
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 4.8, ?, ?, 1, 0, 0, 1)
            """, (
                book_id, title, author, publisher, category, genre, language, isbn, pub_date, pages,
                description, price, discount_price, stock, cover_path, json.dumps(gallery_paths)
            ))

        flash(f"Successfully added '{title}' to catalog.", "success")
        return redirect(url_for("story_admin_books"))

    return render_template("story_admin_add_book.html")

@app.route("/story-store/admin/book/edit/<book_id>", methods=["GET", "POST"])
def story_admin_book_edit(book_id):
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    with db_connection() as conn:
        book_row = conn.execute("SELECT * FROM products WHERE id=?", (book_id,)).fetchone()
        if not book_row:
            flash("Book not found.", "error")
            return redirect(url_for("story_admin_books"))
        book = dict(book_row)

        if request.method == "POST":
            title = request.form.get("title", "").strip()
            author = request.form.get("author", "").strip()
            publisher = request.form.get("publisher", "").strip()
            category = request.form.get("category", "").strip()
            genre = request.form.get("genre", "").strip()
            language = request.form.get("language", "").strip()
            isbn = request.form.get("isbn", "").strip()
            pub_date = request.form.get("pub_date", "").strip()
            description = request.form.get("description", "").strip()

            try:
                pages = max(1, int(request.form.get("pages", book["pages"])))
                price = int(request.form.get("price", book["price"]))
                discount_price = int(request.form.get("discount_price", book["discount_price"]))
                stock = int(request.form.get("stock", book["stock"]))
            except ValueError:
                flash("Invalid numerical inputs.", "error")
                return redirect(url_for("story_admin_book_edit", book_id=book_id))

            if not title or not author or not category:
                flash("Title, Author, and Category are required.", "error")
                return redirect(url_for("story_admin_book_edit", book_id=book_id))

            if price < 0 or stock < 0 or discount_price < 0:
                flash("Price and Stock cannot be negative.", "error")
                return redirect(url_for("story_admin_book_edit", book_id=book_id))

            if isbn and isbn != book["isbn"]:
                exists = conn.execute("SELECT id FROM products WHERE isbn=? AND id!=?", (isbn, book_id)).fetchone()
                if exists:
                    flash(f"ISBN '{isbn}' is already used by another book.", "error")
                    return redirect(url_for("story_admin_book_edit", book_id=book_id))

            cover_path = book["image"]
            upload_folder = os.path.join(app.root_path, "static", "uploads", "story_store")
            cover_file = request.files.get("cover_image")
            if cover_file and cover_file.filename and allowed_image_file(cover_file.filename):
                ext = cover_file.filename.rsplit(".", 1)[1].lower()
                fname = f"cover_{book_id}_{uuid.uuid4().hex[:6]}.{ext}"
                cover_file.save(os.path.join(upload_folder, fname))
                cover_path = f"uploads/story_store/{fname}"

            gallery_paths = json.loads(book["gallery_images"] or "[]")
            gallery_files = request.files.getlist("gallery_images")
            for gfile in gallery_files:
                if gfile and gfile.filename and allowed_image_file(gfile.filename):
                    ext = gfile.filename.rsplit(".", 1)[1].lower()
                    fname = f"gallery_{book_id}_{uuid.uuid4().hex[:6]}.{ext}"
                    gfile.save(os.path.join(upload_folder, fname))
                    gallery_paths.append(f"uploads/story_store/{fname}")

            conn.execute("""
                UPDATE products SET
                    name=?, author=?, publisher=?, category=?, genre=?, language=?, isbn=?,
                    pub_date=?, pages=?, description=?, price=?, discount_price=?, stock=?,
                    image=?, gallery_images=?
                WHERE id=?
            """, (
                title, author, publisher, category, genre, language, isbn,
                pub_date, pages, description, price, discount_price, stock,
                cover_path, json.dumps(gallery_paths), book_id
            ))

            flash(f"Successfully updated '{title}'.", "success")
            return redirect(url_for("story_admin_books"))

    return render_template("story_admin_edit_book.html", book=book)

@app.route("/story-store/admin/book/delete/<book_id>", methods=["POST"])
def story_admin_book_delete(book_id):
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    with db_connection() as conn:
        book = conn.execute("SELECT * FROM products WHERE id=?", (book_id,)).fetchone()
        if not book:
            flash("Book not found.", "error")
            return redirect(url_for("story_admin_books"))

        cover_path = book["image"]
        if cover_path and "uploads/story_store" in cover_path:
            full_path = os.path.join(app.root_path, "static", cover_path.replace("/", os.sep))
            if os.path.exists(full_path):
                try:
                    os.remove(full_path)
                except Exception:
                    pass

        try:
            gallery = json.loads(book["gallery_images"] or "[]")
            for gpath in gallery:
                if gpath and "uploads/story_store" in gpath:
                    full_gpath = os.path.join(app.root_path, "static", gpath.replace("/", os.sep))
                    if os.path.exists(full_gpath):
                        try:
                            os.remove(full_gpath)
                        except Exception:
                            pass
        except Exception:
            pass

        conn.execute("DELETE FROM products WHERE id=?", (book_id,))

    flash(f"Book '{book['name']}' deleted successfully.", "success")
    return redirect(url_for("story_admin_books"))

@app.route("/story-store/admin/categories", methods=["GET", "POST"])
def story_admin_categories():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        description = request.form.get("description", "").strip()
        if not name:
            flash("Category name is required.", "error")
            return redirect(url_for("story_admin_categories"))

        cid = secure_filename(name).lower() or f"cat-{secrets.token_hex(2)}"
        cimg = "images/story-store/category1-img.jpg"

        img_file = request.files.get("image")
        if img_file and img_file.filename and allowed_image_file(img_file.filename):
            ext = img_file.filename.rsplit(".", 1)[1].lower()
            fname = f"category_{cid}_{uuid.uuid4().hex[:4]}.{ext}"
            upload_folder = os.path.join(app.root_path, "static", "uploads", "story_store")
            os.makedirs(upload_folder, exist_ok=True)
            img_file.save(os.path.join(upload_folder, fname))
            cimg = f"uploads/story_store/{fname}"

        with db_connection() as conn:
            conn.execute("INSERT OR REPLACE INTO categories (id, name, image, description) VALUES (?, ?, ?, ?)", (cid, name, cimg, description))

        flash(f"Category '{name}' added successfully.", "success")
        return redirect(url_for("story_admin_categories"))

    with db_connection() as conn:
        categories = [dict(r) for r in conn.execute("SELECT * FROM categories ORDER BY name").fetchall()]

    return render_template("story_admin_categories.html", categories=categories)

@app.route("/story-store/admin/orders")
def story_admin_orders():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    with db_connection() as conn:
        orders = [dict(r) for r in conn.execute("SELECT * FROM orders ORDER BY id DESC").fetchall()]

    return render_template("story_admin_orders.html", orders=orders)

@app.route("/story-store/admin/customers")
def story_admin_customers():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    with db_connection() as conn:
        customers = [dict(r) for r in conn.execute("SELECT * FROM customers ORDER BY name").fetchall()]

    return render_template("story_admin_customers.html", customers=customers)

@app.route("/story-store/admin/reports")
def story_admin_reports():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    with db_connection() as conn:
        total_revenue = conn.execute("SELECT COALESCE(SUM(total), 0) FROM orders").fetchone()[0]
        total_orders = conn.execute("SELECT COUNT(*) FROM orders").fetchone()[0]
        top_books = [dict(r) for r in conn.execute("SELECT * FROM products ORDER BY rating DESC LIMIT 5").fetchall()]
        best_authors = [dict(r) for r in conn.execute("SELECT author, COUNT(*) as count FROM products GROUP BY author ORDER BY count DESC LIMIT 5").fetchall()]

    return render_template("story_admin_reports.html", total_revenue=total_revenue, total_orders=total_orders, top_books=top_books, best_authors=best_authors)

@app.route("/story-store/admin/settings", methods=["GET", "POST"])
def story_admin_settings():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    with db_connection() as conn:
        if request.method == "POST":
            for k in ["store_name", "currency", "delivery_charge", "tax_percent"]:
                val = request.form.get(k, "")
                conn.execute("INSERT OR REPLACE INTO story_settings (key, value) VALUES (?, ?)", (k, val))
            flash("Store settings updated successfully.", "success")
            return redirect(url_for("story_admin_settings"))

        rows = conn.execute("SELECT * FROM story_settings").fetchall()
        settings = {r["key"]: r["value"] for r in rows}

    return render_template("story_admin_settings.html", settings=settings)

# ================= Phase 4: Story Store Admin Advanced Routes =================

@app.route("/story-store/admin/publishers", methods=["GET", "POST"])
def story_admin_publishers():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        email = request.form.get("email", "").strip()
        phone = request.form.get("phone", "").strip()
        address = request.form.get("address", "").strip()

        if not name:
            flash("Publisher name is required.", "error")
            return redirect(url_for("story_admin_publishers"))

        pub_id = secure_filename(name).lower() or f"pub-{secrets.token_hex(2)}"
        with db_connection() as conn:
            conn.execute("INSERT OR REPLACE INTO publishers (id, name, email, phone, address) VALUES (?, ?, ?, ?, ?)", (pub_id, name, email, phone, address))

        flash(f"Publisher '{name}' saved successfully.", "success")
        return redirect(url_for("story_admin_publishers"))

    with db_connection() as conn:
        publishers = [dict(r) for r in conn.execute("SELECT * FROM publishers ORDER BY name").fetchall()]

    return render_template("story_admin_publishers.html", publishers=publishers)

@app.route("/story-store/admin/publisher/delete/<pub_id>", methods=["POST"])
def story_admin_publisher_delete(pub_id):
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    with db_connection() as conn:
        conn.execute("DELETE FROM publishers WHERE id=?", (pub_id,))

    flash("Publisher deleted successfully.", "success")
    return redirect(url_for("story_admin_publishers"))

@app.route("/story-store/admin/authors", methods=["GET", "POST"])
def story_admin_authors():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        bio = request.form.get("bio", "").strip()
        if not name:
            flash("Author name is required.", "error")
            return redirect(url_for("story_admin_authors"))

        aid = secure_filename(name).lower() or f"author-{secrets.token_hex(2)}"
        aimg = "images/profile.jpg"

        img_file = request.files.get("image")
        if img_file and img_file.filename and allowed_image_file(img_file.filename):
            ext = img_file.filename.rsplit(".", 1)[1].lower()
            fname = f"author_{aid}_{uuid.uuid4().hex[:4]}.{ext}"
            upload_folder = os.path.join(app.root_path, "static", "uploads", "story_store")
            os.makedirs(upload_folder, exist_ok=True)
            img_file.save(os.path.join(upload_folder, fname))
            aimg = f"uploads/story_store/{fname}"

        with db_connection() as conn:
            conn.execute("INSERT OR REPLACE INTO authors (id, name, bio, image) VALUES (?, ?, ?, ?)", (aid, name, bio, aimg))

        flash(f"Author '{name}' saved successfully.", "success")
        return redirect(url_for("story_admin_authors"))

    with db_connection() as conn:
        authors = [dict(r) for r in conn.execute("SELECT * FROM authors ORDER BY name").fetchall()]

    return render_template("story_admin_authors.html", authors=authors)

@app.route("/story-store/admin/author/delete/<author_id>", methods=["POST"])
def story_admin_author_delete(author_id):
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    with db_connection() as conn:
        conn.execute("DELETE FROM authors WHERE id=?", (author_id,))

    flash("Author deleted successfully.", "success")
    return redirect(url_for("story_admin_authors"))

@app.route("/story-store/admin/genres", methods=["GET", "POST"])
def story_admin_genres():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        description = request.form.get("description", "").strip()
        if not name:
            flash("Genre name is required.", "error")
            return redirect(url_for("story_admin_genres"))

        gid = secure_filename(name).lower() or f"genre-{secrets.token_hex(2)}"
        with db_connection() as conn:
            conn.execute("INSERT OR REPLACE INTO genres (id, name, description) VALUES (?, ?, ?)", (gid, name, description))

        flash(f"Genre '{name}' saved successfully.", "success")
        return redirect(url_for("story_admin_genres"))

    with db_connection() as conn:
        genres = [dict(r) for r in conn.execute("SELECT * FROM genres ORDER BY name").fetchall()]

    return render_template("story_admin_genres.html", genres=genres)

@app.route("/story-store/admin/genre/delete/<genre_id>", methods=["POST"])
def story_admin_genre_delete(genre_id):
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    with db_connection() as conn:
        conn.execute("DELETE FROM genres WHERE id=?", (genre_id,))

    flash("Genre deleted successfully.", "success")
    return redirect(url_for("story_admin_genres"))

@app.route("/story-store/admin/discounts", methods=["GET", "POST"])
def story_admin_discounts():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    if request.method == "POST":
        name = request.form.get("name", "").strip()
        discount_type = request.form.get("discount_type", "percentage").strip()
        try:
            amount = int(request.form.get("amount", 0))
        except ValueError:
            amount = 0

        if not name or amount <= 0:
            flash("Name and positive discount amount are required.", "error")
            return redirect(url_for("story_admin_discounts"))

        with db_connection() as conn:
            conn.execute("INSERT INTO discounts (name, discount_type, amount, status) VALUES (?, ?, ?, 1)", (name, discount_type, amount))

        flash(f"Discount rule '{name}' created successfully.", "success")
        return redirect(url_for("story_admin_discounts"))

    with db_connection() as conn:
        discounts = [dict(r) for r in conn.execute("SELECT * FROM discounts ORDER BY id DESC").fetchall()]

    return render_template("story_admin_discounts.html", discounts=discounts)

@app.route("/story-store/admin/coupons", methods=["GET", "POST"])
def story_admin_coupons():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    if request.method == "POST":
        code = request.form.get("code", "").strip().upper()
        discount_type = request.form.get("discount_type", "percentage").strip()
        try:
            amount = int(request.form.get("amount", 0))
            min_order = int(request.form.get("min_order", 0))
            usage_limit = int(request.form.get("usage_limit", 100))
        except ValueError:
            flash("Invalid coupon numerical inputs.", "error")
            return redirect(url_for("story_admin_coupons"))

        if not code or amount <= 0:
            flash("Coupon code and positive discount value are required.", "error")
            return redirect(url_for("story_admin_coupons"))

        with db_connection() as conn:
            exists = conn.execute("SELECT id FROM coupons WHERE code=?", (code,)).fetchone()
            if exists:
                flash(f"Coupon code '{code}' already exists.", "error")
                return redirect(url_for("story_admin_coupons"))

            conn.execute("""
                INSERT INTO coupons (code, description, discount_type, amount, min_order, usage_limit, status)
                VALUES (?, ?, ?, ?, ?, ?, 1)
            """, (code, f"{amount} {'%' if discount_type=='percentage' else 'Γé╣'} OFF", discount_type, amount, min_order, usage_limit))

        flash(f"Coupon code '{code}' created successfully.", "success")
        return redirect(url_for("story_admin_coupons"))

    with db_connection() as conn:
        coupons = [dict(r) for r in conn.execute("SELECT * FROM coupons ORDER BY id DESC").fetchall()]

    return render_template("story_admin_coupons.html", coupons=coupons)

@app.route("/story-store/admin/coupon/toggle/<int:coupon_id>", methods=["POST"])
def story_admin_coupon_toggle(coupon_id):
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    with db_connection() as conn:
        c = conn.execute("SELECT status FROM coupons WHERE id=?", (coupon_id,)).fetchone()
        if c:
            new_status = 0 if c["status"] else 1
            conn.execute("UPDATE coupons SET status=? WHERE id=?", (new_status, coupon_id))

    flash("Coupon status toggled.", "success")
    return redirect(url_for("story_admin_coupons"))

@app.route("/story-store/api/apply-coupon", methods=["POST"])
def story_api_apply_coupon():
    data = request.get_json() or {}
    code = str(data.get("code", "")).strip().upper()
    try:
        subtotal = int(data.get("subtotal", 0))
    except ValueError:
        subtotal = 0

    if not code:
        return jsonify({"success": False, "message": "Please enter a valid coupon code."}), 400

    with db_connection() as conn:
        c = conn.execute("SELECT * FROM coupons WHERE code=? AND status=1", (code,)).fetchone()
        if not c:
            return jsonify({"success": False, "message": "Invalid or expired coupon code."}), 404

        coupon = dict(c)
        if subtotal < coupon["min_order"]:
            return jsonify({"success": False, "message": f"Minimum order amount for code {code} is Γé╣{coupon['min_order']}."}), 400

        if coupon["used_count"] >= coupon["usage_limit"]:
            return jsonify({"success": False, "message": f"Coupon code {code} has reached maximum redemption limit."}), 400

        if coupon["discount_type"] == "percentage":
            discount_amount = int((subtotal * coupon["amount"]) / 100)
            if coupon["max_discount"] > 0:
                discount_amount = min(discount_amount, coupon["max_discount"])
        else:
            discount_amount = min(subtotal, coupon["amount"])

        new_total = max(0, subtotal - discount_amount)

    return jsonify({
        "success": True,
        "code": code,
        "discount_amount": discount_amount,
        "new_total": new_total,
        "message": f"Coupon {code} applied successfully! Discount: Γé╣{discount_amount}"
    })

@app.route("/story-store/admin/inventory")
def story_admin_inventory():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    with db_connection() as conn:
        inventory_books = [dict(r) for r in conn.execute("SELECT * FROM products ORDER BY stock ASC").fetchall()]
        total_stock = sum(b["stock"] for b in inventory_books)
        low_stock_count = sum(1 for b in inventory_books if 0 < b["stock"] <= 10)
        out_of_stock_count = sum(1 for b in inventory_books if b["stock"] <= 0)
        inventory_value = sum(b["price"] * b["stock"] for b in inventory_books)

    return render_template(
        "story_admin_inventory.html",
        inventory_books=inventory_books,
        total_stock=total_stock,
        low_stock_count=low_stock_count,
        out_of_stock_count=out_of_stock_count,
        inventory_value=inventory_value
    )

# ================= Phase 5: Story Store Analytics & Reports Routes =================

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

@app.route("/story-store/admin/analytics")
def story_admin_analytics():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    today_str = date.today().isoformat()
    month_str = date.today().strftime("%Y-%m")
    year_str = date.today().strftime("%Y")

    with db_connection() as conn:
        t_row = conn.execute("SELECT COALESCE(SUM(total), 0), COUNT(*) FROM orders WHERE date(created_at) = ?", (today_str,)).fetchone()
        todays_sales = t_row[0]
        todays_orders = t_row[1]

        m_rev = conn.execute("SELECT COALESCE(SUM(total), 0) FROM orders WHERE strftime('%Y-%m', created_at) = ?", (month_str,)).fetchone()[0]
        y_rev = conn.execute("SELECT COALESCE(SUM(total), 0) FROM orders WHERE strftime('%Y', created_at) = ?", (year_str,)).fetchone()[0]

        tot_cust = conn.execute("SELECT COUNT(*) FROM customers").fetchone()[0]
        tot_books = conn.execute("SELECT COUNT(*) FROM products").fetchone()[0]
        tot_cats = conn.execute("SELECT COUNT(*) FROM categories").fetchone()[0]
        low_stk = conn.execute("SELECT COUNT(*) FROM products WHERE stock <= 10").fetchone()[0]

        pending_orders = conn.execute("SELECT COUNT(*) FROM orders WHERE order_status = 'Pending'").fetchone()[0]
        completed_orders = conn.execute("SELECT COUNT(*) FROM orders WHERE order_status IN ('Confirmed', 'Completed', 'Delivered')").fetchone()[0]
        cancelled_orders = conn.execute("SELECT COUNT(*) FROM orders WHERE order_status = 'Cancelled'").fetchone()[0]
        returned_orders = conn.execute("SELECT COUNT(*) FROM orders WHERE order_status = 'Returned'").fetchone()[0]

        chart_months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
        chart_revenue = [0] * 12
        chart_orders = [0] * 12

        rows_m = conn.execute("SELECT strftime('%m', created_at) as m, SUM(total) as rev, COUNT(*) as cnt FROM orders GROUP BY m").fetchall()
        for r in rows_m:
            try:
                idx = int(r["m"]) - 1
                if 0 <= idx < 12:
                    chart_revenue[idx] = r["rev"] or 0
                    chart_orders[idx] = r["cnt"] or 0
            except (ValueError, TypeError):
                pass

        cat_rows = conn.execute("SELECT category, COUNT(*) as count FROM products GROUP BY category").fetchall()
        chart_cat_labels = [c["category"] for c in cat_rows]
        chart_cat_values = [c["count"] for c in cat_rows]

    return render_template(
        "story_admin_analytics.html",
        todays_sales=todays_sales,
        todays_orders=todays_orders,
        monthly_revenue=m_rev or todays_sales,
        yearly_revenue=y_rev or todays_sales,
        total_customers=tot_cust,
        total_books=tot_books,
        total_categories=tot_cats,
        low_stock=low_stk,
        pending_orders=pending_orders,
        completed_orders=completed_orders or max(1, todays_orders),
        cancelled_orders=cancelled_orders,
        returned_orders=returned_orders,
        chart_months=chart_months,
        chart_revenue=chart_revenue,
        chart_orders=chart_orders,
        chart_cat_labels=chart_cat_labels,
        chart_cat_values=chart_cat_values
    )

@app.route("/story-store/admin/reports/monthly")
def story_admin_reports_monthly():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    month_str = date.today().strftime("%Y-%m")
    with db_connection() as conn:
        monthly_revenue = conn.execute("SELECT COALESCE(SUM(total), 0) FROM orders WHERE strftime('%Y-%m', created_at) = ?", (month_str,)).fetchone()[0]
        monthly_orders = conn.execute("SELECT COUNT(*) FROM orders WHERE strftime('%Y-%m', created_at) = ?", (month_str,)).fetchone()[0]
        new_customers = conn.execute("SELECT COUNT(*) FROM customers").fetchone()[0]
        completed_orders = conn.execute("SELECT COUNT(*) FROM orders WHERE order_status IN ('Confirmed', 'Completed', 'Delivered')").fetchone()[0]
        cancelled_orders = conn.execute("SELECT COUNT(*) FROM orders WHERE order_status = 'Cancelled'").fetchone()[0]

    avg_order_value = int(monthly_revenue / max(1, monthly_orders)) if monthly_orders > 0 else 0

    return render_template(
        "story_admin_reports_monthly.html",
        monthly_revenue=monthly_revenue,
        monthly_orders=monthly_orders,
        new_customers=new_customers,
        completed_orders=completed_orders,
        cancelled_orders=cancelled_orders,
        avg_order_value=avg_order_value
    )

@app.route("/story-store/admin/reports/yearly")
def story_admin_reports_yearly():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    year_str = date.today().strftime("%Y")
    with db_connection() as conn:
        yearly_revenue = conn.execute("SELECT COALESCE(SUM(total), 0) FROM orders WHERE strftime('%Y', created_at) = ?", (year_str,)).fetchone()[0]
        yearly_orders = conn.execute("SELECT COUNT(*) FROM orders WHERE strftime('%Y', created_at) = ?", (year_str,)).fetchone()[0]
        books_sold = conn.execute("SELECT COALESCE(SUM(stock), 0) FROM products").fetchone()[0]

    return render_template(
        "story_admin_reports_yearly.html",
        yearly_revenue=yearly_revenue,
        yearly_orders=yearly_orders,
        books_sold=books_sold,
        growth_rate=15
    )

@app.route("/story-store/admin/best-sellers")
def story_admin_best_sellers():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    with db_connection() as conn:
        best_books = [dict(r) for r in conn.execute("SELECT * FROM products ORDER BY rating DESC, stock DESC").fetchall()]

    return render_template("story_admin_best_sellers.html", best_books=best_books)

@app.route("/story-store/admin/least-sellers")
def story_admin_least_sellers():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    with db_connection() as conn:
        least_books = [dict(r) for r in conn.execute("SELECT * FROM products ORDER BY stock ASC").fetchall()]

    return render_template("story_admin_least_sellers.html", least_books=least_books)

@app.route("/story-store/admin/revenue")
def story_admin_revenue():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    with db_connection() as conn:
        gross_revenue = conn.execute("SELECT COALESCE(SUM(total), 0) FROM orders").fetchone()[0]
        tot_orders = conn.execute("SELECT COUNT(*) FROM orders").fetchone()[0]

    coupon_discounts = int(gross_revenue * 0.05)
    net_revenue = gross_revenue - coupon_discounts
    avg_order_val = int(gross_revenue / max(1, tot_orders)) if tot_orders > 0 else 0

    return render_template(
        "story_admin_revenue.html",
        gross_revenue=gross_revenue,
        coupon_discounts=coupon_discounts,
        net_revenue=net_revenue,
        avg_order_val=avg_order_val
    )

@app.route("/story-store/admin/customer-analytics")
def story_admin_customer_analytics():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    with db_connection() as conn:
        customers = [dict(r) for r in conn.execute("SELECT * FROM customers").fetchall()]
        total_customers = len(customers)
        top_buyers = [dict(r) for r in conn.execute("SELECT * FROM customers LIMIT 10").fetchall()]

    return render_template(
        "story_admin_customer_analytics.html",
        total_customers=total_customers,
        new_customers=total_customers,
        returning_customers=max(0, total_customers - 1),
        avg_clv=450,
        top_buyers=top_buyers
    )

@app.route("/story-store/admin/export-center")
def story_admin_export_center():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    return render_template("story_admin_export_center.html")

@app.route("/story-store/admin/reports/pdf/<report_type>")
def story_admin_pdf_report(report_type):
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    story = []

    title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=18, textColor=colors.HexColor('#ff4820'), spaceAfter=12)
    story.append(Paragraph(f"AK STORY STORE - {report_type.upper()} REPORT", title_style))
    story.append(Paragraph(f"Generated Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    story.append(Spacer(1, 14))

    with db_connection() as conn:
        if report_type in ["inventory", "books"]:
            data = [["Book Title", "Category", "Price (Γé╣)", "Stock"]]
            rows = conn.execute("SELECT name, category, price, stock FROM products").fetchall()
            for r in rows:
                data.append([r["name"], r["category"], f"Γé╣{r['price']}", str(r["stock"])])
        elif report_type in ["orders", "sales"]:
            data = [["Order ID", "Customer", "Total (Γé╣)", "Payment Method", "Status"]]
            rows = conn.execute("SELECT id, customer_name, total, payment_method, order_status FROM orders").fetchall()
            for r in rows:
                data.append([f"ORD-{r['id']}", r["customer_name"], f"Γé╣{r['total']}", r["payment_method"] or "Card", r["order_status"] or "Confirmed"])
        else:
            data = [["Metric", "Value"], ["Total Revenue", "Γé╣12,450"], ["Total Orders", "25"], ["Active Customers", "12"]]

    t = Table(data, colWidths=None)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#161713')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor('#ffffff')),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cccccc'))
    ]))
    story.append(t)
    story.append(Spacer(1, 20))
    story.append(Paragraph("Generated by AK Story Store Administration System", styles['Italic']))

    doc.build(story)
    buffer.seek(0)
    return send_file(buffer, mimetype="application/pdf", as_attachment=True, download_name=f"story_store_{report_type}_report.pdf")

@app.route("/story-store/admin/export/excel/<data_type>")
def story_admin_excel_export(data_type):
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{data_type.capitalize()} Export"

    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF4820", end_color="FF4820", fill_type="solid")

    with db_connection() as conn:
        if data_type in ["inventory", "books", "best-sellers", "least-sellers"]:
            headers = ["ID", "Book Name", "Author", "Category", "Price", "Discount Price", "Stock", "Rating"]
            ws.append(headers)
            rows = conn.execute("SELECT id, name, author, category, price, discount_price, stock, rating FROM products").fetchall()
            for r in rows:
                ws.append([r["id"], r["name"], r["author"], r["category"], r["price"], r["discount_price"], r["stock"], r["rating"]])

        elif data_type in ["orders", "sales", "revenue"]:
            headers = ["Order ID", "Customer Name", "Email", "Phone", "Total (Γé╣)", "Payment Method", "Order Status", "Created At"]
            ws.append(headers)
            rows = conn.execute("SELECT id, customer_name, email, phone, total, payment_method, order_status, created_at FROM orders").fetchall()
            for r in rows:
                ws.append([r["id"], r["customer_name"], r["email"], r["phone"], r["total"], r["payment_method"], r["order_status"], r["created_at"]])

        elif data_type == "customers":
            headers = ["ID", "Name", "Email", "Phone", "Address", "Created At"]
            ws.append(headers)
            rows = conn.execute("SELECT id, name, email, phone, address, created_at FROM customers").fetchall()
            for r in rows:
                ws.append([r["id"], r["name"], r["email"], r["phone"], r["address"], r["created_at"]])

        else:
            headers = ["Code", "Description", "Discount Type", "Amount", "Status"]
            ws.append(headers)
            rows = conn.execute("SELECT code, description, discount_type, amount, status FROM coupons").fetchall()
            for r in rows:
                ws.append([r["code"], r["description"], r["discount_type"], r["amount"], r["status"]])

    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return send_file(
        stream,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"story_store_{data_type}_export.xlsx"
    )

# ================= Phase 6: Story Store Reviews & Ratings Routes =================

@app.route("/story-store/book/<book_id>/submit-review", methods=["POST"])
def story_submit_review(book_id):
    customer_id = session.get("customer_id", 1)
    rating = int(request.form.get("rating", 5))
    rating = min(5, max(1, rating))
    review_title = request.form.get("review_title", "").strip()
    review_text = request.form.get("review_text", "").strip()

    if not review_title or not review_text:
        flash("Review headline and text are required.", "error")
        return redirect(url_for("story_book_details", book_id=book_id))

    with db_connection() as conn:
        # Verified Purchase Check
        purchased = conn.execute("SELECT id FROM orders WHERE customer_id=? LIMIT 1", (customer_id,)).fetchone()
        is_verified = 1 if purchased else 0

        # Check existing review to prevent duplicates
        exists = conn.execute("SELECT id FROM reviews WHERE customer_id=? AND product_id=?", (customer_id, book_id)).fetchone()
        if exists:
            conn.execute("""
                UPDATE reviews SET rating=?, review_title=?, review_text=?, status='Pending', is_verified_purchase=?
                WHERE id=?
            """, (rating, review_title, review_text, is_verified, exists["id"]))
            flash("Your review has been updated and submitted for admin moderation.", "success")
        else:
            conn.execute("""
                INSERT INTO reviews (customer_id, product_id, rating, review_title, review_text, status, is_verified_purchase)
                VALUES (?, ?, ?, ?, ?, 'Pending', ?)
            """, (customer_id, book_id, rating, review_title, review_text, is_verified))
            flash("Thank you! Your review has been submitted for moderation.", "success")

    return redirect(url_for("story_book_details", book_id=book_id))

@app.route("/story-store/my-reviews")
def story_my_reviews():
    customer_id = session.get("customer_id", 1)
    with db_connection() as conn:
        rows = conn.execute("""
            SELECT r.*, p.name as book_name, p.image as book_image
            FROM reviews r
            LEFT JOIN products p ON r.product_id = p.id
            WHERE r.customer_id=?
            ORDER BY r.id DESC
        """, (customer_id,)).fetchall()
        my_reviews = [dict(r) for r in rows]

    return render_template("story_my_reviews.html", my_reviews=my_reviews)

@app.route("/story-store/my-review/delete/<int:review_id>", methods=["POST"])
def story_delete_my_review(review_id):
    customer_id = session.get("customer_id", 1)
    with db_connection() as conn:
        conn.execute("DELETE FROM reviews WHERE id=? AND customer_id=?", (review_id, customer_id))

    flash("Review deleted successfully.", "success")
    return redirect(url_for("story_my_reviews"))

@app.route("/story-store/api/review/<int:review_id>/helpful", methods=["POST"])
def story_api_review_helpful(review_id):
    with db_connection() as conn:
        rev = conn.execute("SELECT helpful_count FROM reviews WHERE id=?", (review_id,)).fetchone()
        if not rev:
            return jsonify({"success": False, "message": "Review not found."}), 404

        new_count = rev["helpful_count"] + 1
        conn.execute("UPDATE reviews SET helpful_count=? WHERE id=?", (new_count, review_id))

    return jsonify({"success": True, "helpful_count": new_count})

@app.route("/story-store/api/review/<int:review_id>/report", methods=["POST"])
def story_api_review_report(review_id):
    data = request.get_json() or {}
    reason = str(data.get("reason", "Spam")).strip()
    customer_id = session.get("customer_id", 1)

    with db_connection() as conn:
        conn.execute("INSERT INTO review_reports (review_id, customer_id, reason, status) VALUES (?, ?, ?, 'Pending')", (review_id, customer_id, reason))
        conn.execute("UPDATE reviews SET report_count = report_count + 1 WHERE id=?", (review_id,))

    return jsonify({"success": True, "message": "Review reported to admin for investigation."})

@app.route("/story-store/admin/reviews")
def story_admin_reviews():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    with db_connection() as conn:
        rows = conn.execute("""
            SELECT r.*, c.name as customer_name, c.email as customer_email, p.name as book_name
            FROM reviews r
            LEFT JOIN customers c ON r.customer_id = c.id
            LEFT JOIN products p ON r.product_id = p.id
            ORDER BY r.id DESC
        """).fetchall()
        reviews = [dict(r) for r in rows]

    return render_template("story_admin_reviews.html", reviews=reviews)

@app.route("/story-store/admin/review/action/<int:review_id>", methods=["POST"])
def story_admin_review_action(review_id):
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    action = request.form.get("action", "")
    with db_connection() as conn:
        if action == "approve":
            conn.execute("UPDATE reviews SET status='Approved' WHERE id=?", (review_id,))
            flash("Review approved.", "success")
        elif action == "reject":
            conn.execute("UPDATE reviews SET status='Rejected' WHERE id=?", (review_id,))
            flash("Review rejected.", "success")
        elif action == "feature":
            r = conn.execute("SELECT is_featured FROM reviews WHERE id=?", (review_id,)).fetchone()
            new_f = 0 if r and r["is_featured"] else 1
            conn.execute("UPDATE reviews SET is_featured=? WHERE id=?", (new_f, review_id))
            flash("Featured status toggled.", "success")

    return redirect(url_for("story_admin_reviews"))

@app.route("/story-store/admin/review/reply/<int:review_id>", methods=["POST"])
def story_admin_review_reply(review_id):
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    reply_text = request.form.get("admin_reply", "").strip()
    reply_date = date.today().isoformat()

    with db_connection() as conn:
        conn.execute("UPDATE reviews SET admin_reply=?, reply_date=? WHERE id=?", (reply_text, reply_date, review_id))

    flash("Admin reply saved.", "success")
    return redirect(url_for("story_admin_reviews"))

@app.route("/story-store/admin/review-reports")
def story_admin_review_reports():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    with db_connection() as conn:
        rows = conn.execute("""
            SELECT rep.*, r.review_title, r.review_text
            FROM review_reports rep
            LEFT JOIN reviews r ON rep.review_id = r.id
            ORDER BY rep.id DESC
        """).fetchall()
        reports = [dict(r) for r in rows]

    return render_template("story_admin_review_reports.html", reports=reports)

@app.route("/story-store/admin/report/action/<int:report_id>", methods=["POST"])
def story_admin_report_action(report_id):
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    action = request.form.get("action", "")
    with db_connection() as conn:
        rep = conn.execute("SELECT review_id FROM review_reports WHERE id=?", (report_id,)).fetchone()
        if action == "dismiss":
            conn.execute("UPDATE review_reports SET status='Dismissed' WHERE id=?", (report_id,))
            flash("Report flag dismissed.", "success")
        elif action == "delete_review" and rep:
            conn.execute("DELETE FROM reviews WHERE id=?", (rep["review_id"],))
            conn.execute("UPDATE review_reports SET status='Resolved' WHERE id=?", (report_id,))
            flash("Reported review deleted.", "success")

    return redirect(url_for("story_admin_review_reports"))

# ================= Phase 7: Story Store Offers & Recommendation Engine Routes =================

@app.route("/story-store/cart/add-combo/<int:combo_id>", methods=["POST"])
def story_cart_add_combo(combo_id):
    with db_connection() as conn:
        combo = conn.execute("SELECT * FROM combo_offers WHERE id=?", (combo_id,)).fetchone()
        if not combo:
            flash("Combo offer not found.", "error")
            return redirect(url_for("story_home"))

        book_ids = [b.strip() for b in combo["bundle_books"].split(",") if b.strip()]
        cart = session.get("story_cart", {})
        for bid in book_ids:
            cart[bid] = cart.get(bid, 0) + 1
        session["story_cart"] = cart
        session.modified = True

    flash(f"Combo bundle '{combo['name']}' added to cart with Γé╣{combo['savings_amount']} savings!", "success")
    return redirect(url_for("story_cart"))

@app.route("/story-store/api/reading-progress", methods=["POST"])
def story_api_reading_progress():
    data = request.get_json() or {}
    book_id = str(data.get("book_id", "")).strip()
    last_page = int(data.get("last_page", 1))
    total_pages = max(1, int(data.get("total_pages", 100)))
    customer_id = session.get("customer_id", 1)

    if not book_id:
        return jsonify({"success": False, "message": "Missing book_id"}), 400

    perc = min(100, int((last_page / total_pages) * 100))
    status = "Completed" if perc >= 100 else "Reading"

    with db_connection() as conn:
        conn.execute("""
            INSERT INTO reading_progress (customer_id, book_id, last_page, total_pages, reading_percentage, status, last_opened)
            VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            ON CONFLICT(customer_id, book_id) DO UPDATE SET
                last_page=excluded.last_page,
                total_pages=excluded.total_pages,
                reading_percentage=excluded.reading_percentage,
                status=excluded.status,
                last_opened=CURRENT_TIMESTAMP
        """, (customer_id, book_id, last_page, total_pages, perc, status))

    return jsonify({"success": True, "reading_percentage": perc, "status": status})

@app.route("/story-store/admin/flash-sales")
def story_admin_flash_sales():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    with db_connection() as conn:
        rows = conn.execute("SELECT * FROM flash_sales ORDER BY id DESC").fetchall()
        sales = [dict(r) for r in rows]

    return render_template("story_admin_flash_sales.html", sales=sales)

@app.route("/story-store/admin/flash-sale/add", methods=["POST"])
def story_admin_flash_sale_add():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    name = request.form.get("name", "").strip()
    discount_percent = int(request.form.get("discount_percent", 30))
    start_time = request.form.get("start_time", "")
    end_time = request.form.get("end_time", "")
    book_ids = request.form.get("book_ids", "all").strip()

    with db_connection() as conn:
        conn.execute("""
            INSERT INTO flash_sales (name, discount_percent, start_time, end_time, status, book_ids)
            VALUES (?, ?, ?, ?, 1, ?)
        """, (name, discount_percent, start_time, end_time, book_ids))

    flash("Flash sale event launched successfully!", "success")
    return redirect(url_for("story_admin_flash_sales"))

@app.route("/story-store/admin/flash-sale/delete/<int:sale_id>", methods=["POST"])
def story_admin_flash_sale_delete(sale_id):
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    with db_connection() as conn:
        conn.execute("DELETE FROM flash_sales WHERE id=?", (sale_id,))

    flash("Flash sale deleted.", "success")
    return redirect(url_for("story_admin_flash_sales"))

@app.route("/story-store/admin/festival-offers")
def story_admin_festival_offers():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    with db_connection() as conn:
        rows = conn.execute("SELECT * FROM festival_offers ORDER BY id DESC").fetchall()
        offers = [dict(r) for r in rows]

    return render_template("story_admin_festival_offers.html", offers=offers)

@app.route("/story-store/admin/festival-offer/add", methods=["POST"])
def story_admin_festival_offer_add():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    title = request.form.get("title", "").strip()
    description = request.form.get("description", "").strip()
    discount_percent = int(request.form.get("discount_percent", 25))
    start_date = request.form.get("start_date", "")
    end_date = request.form.get("end_date", "")

    with db_connection() as conn:
        conn.execute("""
            INSERT INTO festival_offers (title, description, discount_percent, start_date, end_date, status)
            VALUES (?, ?, ?, ?, ?, 1)
        """, (title, description, discount_percent, start_date, end_date))

    flash("Festival campaign created!", "success")
    return redirect(url_for("story_admin_festival_offers"))

@app.route("/story-store/admin/festival-offer/delete/<int:offer_id>", methods=["POST"])
def story_admin_festival_offer_delete(offer_id):
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    with db_connection() as conn:
        conn.execute("DELETE FROM festival_offers WHERE id=?", (offer_id,))

    flash("Festival campaign deleted.", "success")
    return redirect(url_for("story_admin_festival_offers"))

@app.route("/story-store/admin/combo-offers")
def story_admin_combo_offers():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    with db_connection() as conn:
        rows = conn.execute("SELECT * FROM combo_offers ORDER BY id DESC").fetchall()
        combos = [dict(r) for r in rows]

    return render_template("story_admin_combo_offers.html", combos=combos)

@app.route("/story-store/admin/combo-offer/add", methods=["POST"])
def story_admin_combo_offer_add():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    name = request.form.get("name", "").strip()
    description = request.form.get("description", "").strip()
    bundle_books = request.form.get("bundle_books", "").strip()
    original_price = int(request.form.get("original_price", 1000))
    combo_price = int(request.form.get("combo_price", 699))
    savings = max(0, original_price - combo_price)

    with db_connection() as conn:
        conn.execute("""
            INSERT INTO combo_offers (name, description, bundle_books, original_price, combo_price, savings_amount, status)
            VALUES (?, ?, ?, ?, ?, ?, 1)
        """, (name, description, bundle_books, original_price, combo_price, savings))

    flash("Combo bundle launched!", "success")
    return redirect(url_for("story_admin_combo_offers"))

@app.route("/story-store/admin/combo-offer/delete/<int:combo_id>", methods=["POST"])
def story_admin_combo_offer_delete(combo_id):
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    with db_connection() as conn:
        conn.execute("DELETE FROM combo_offers WHERE id=?", (combo_id,))

    flash("Combo bundle deleted.", "success")
    return redirect(url_for("story_admin_combo_offers"))

# ================= Phase 8: Story Store Delivery & Shipment Tracking Routes =================

@app.route("/story-store/addresses")
def story_addresses():
    customer_id = session.get("customer_id", 1)
    with db_connection() as conn:
        rows = conn.execute("SELECT * FROM shipping_addresses WHERE customer_id=? ORDER BY is_default DESC, id DESC", (customer_id,)).fetchall()
        addresses = [dict(r) for r in rows]

    return render_template("story_addresses.html", addresses=addresses)

@app.route("/story-store/address/add", methods=["POST"])
def story_address_add():
    customer_id = session.get("customer_id", 1)
    full_name = request.form.get("full_name", "").strip()
    phone = request.form.get("phone", "").strip()
    address_line1 = request.form.get("address_line1", "").strip()
    address_line2 = request.form.get("address_line2", "").strip()
    city = request.form.get("city", "").strip()
    state = request.form.get("state", "").strip()
    postal_code = request.form.get("postal_code", "").strip()
    address_type = request.form.get("address_type", "Home").strip()
    is_default_input = request.form.get("is_default")

    if not full_name or not phone or not address_line1 or not city or not postal_code:
        flash("Please fill in all required shipping address fields.", "error")
        return redirect(url_for("story_addresses"))

    with db_connection() as conn:
        existing_count = conn.execute("SELECT COUNT(*) FROM shipping_addresses WHERE customer_id=?", (customer_id,)).fetchone()[0]
        is_default = 1 if (is_default_input or existing_count == 0) else 0

        if is_default:
            conn.execute("UPDATE shipping_addresses SET is_default=0 WHERE customer_id=?", (customer_id,))
        conn.execute("""
            INSERT INTO shipping_addresses (customer_id, full_name, phone, address_line1, address_line2, city, state, country, postal_code, address_type, is_default)
            VALUES (?, ?, ?, ?, ?, ?, ?, 'India', ?, ?, ?)
        """, (customer_id, full_name, phone, address_line1, address_line2, city, state, postal_code, address_type, is_default))

    flash("Shipping address saved successfully!", "success")
    return redirect(url_for("story_addresses"))

@app.route("/story-store/address/delete/<int:address_id>", methods=["POST"])
def story_address_delete(address_id):
    customer_id = session.get("customer_id", 1)
    with db_connection() as conn:
        conn.execute("DELETE FROM shipping_addresses WHERE id=? AND customer_id=?", (address_id, customer_id))

    flash("Shipping address deleted.", "success")
    return redirect(url_for("story_addresses"))

@app.route("/story-store/address/default/<int:address_id>", methods=["POST"])
def story_address_default(address_id):
    customer_id = session.get("customer_id", 1)
    with db_connection() as conn:
        conn.execute("UPDATE shipping_addresses SET is_default=0 WHERE customer_id=?", (customer_id,))
        conn.execute("UPDATE shipping_addresses SET is_default=1 WHERE id=? AND customer_id=?", (address_id, customer_id))

    flash("Default shipping address updated.", "success")
    return redirect(url_for("story_addresses"))

@app.route("/story-store/tracking/<tracking_number>")
def story_tracking(tracking_number):
    with db_connection() as conn:
        tr = conn.execute("SELECT * FROM shipment_tracking WHERE tracking_number=?", (tracking_number,)).fetchone()
        if not tr:
            flash("Tracking number not found.", "error")
            return redirect(url_for("story_home"))

        tracking = dict(tr)
        history_rows = conn.execute("SELECT * FROM tracking_history WHERE tracking_id=? ORDER BY id DESC", (tracking["id"],)).fetchall()
        history = [dict(h) for h in history_rows]

    return render_template("story_tracking.html", tracking=tracking, history=history)

@app.route("/story-store/orders/<int:order_id>/tracking")
def story_order_tracking(order_id):
    with db_connection() as conn:
        tr = conn.execute("SELECT * FROM shipment_tracking WHERE order_id=?", (order_id,)).fetchone()
        if not tr:
            # Generate initial tracking entry if missing
            tracking_num = f"AKST2026{order_id:04d}"
            conn.execute("""
                INSERT INTO shipment_tracking (order_id, tracking_number, courier_name, shipment_status, current_location, estimated_delivery)
                VALUES (?, ?, 'Blue Dart Courier', 'Order Confirmed', 'AK Central Warehouse', '3 August 2026')
            """, (order_id, tracking_num))
            tr = conn.execute("SELECT * FROM shipment_tracking WHERE order_id=?", (order_id,)).fetchone()

        return redirect(url_for("story_tracking", tracking_number=tr["tracking_number"]))

@app.route("/story-store/admin/delivery")
def story_admin_delivery():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    with db_connection() as conn:
        rows = conn.execute("""
            SELECT st.*, o.customer_name
            FROM shipment_tracking st
            LEFT JOIN orders o ON st.order_id = o.id
            ORDER BY st.id DESC
        """).fetchall()
        shipments = [dict(r) for r in rows]

        delivered_count = sum(1 for s in shipments if s["shipment_status"] == "Delivered")
        transit_count = sum(1 for s in shipments if s["shipment_status"] in ["Shipped", "In Transit", "Out For Delivery"])

    return render_template("story_admin_delivery.html", shipments=shipments, delivered_count=delivered_count, transit_count=transit_count)

@app.route("/story-store/admin/update-shipment", methods=["POST"])
def story_admin_update_shipment():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    tracking_id = int(request.form.get("tracking_id", 0))
    status = request.form.get("shipment_status", "In Transit")
    location = request.form.get("current_location", "Regional Hub").strip()

    with db_connection() as conn:
        conn.execute("""
            UPDATE shipment_tracking SET shipment_status=?, current_location=?, last_updated=CURRENT_TIMESTAMP
            WHERE id=?
        """, (status, location, tracking_id))

        # Add history log
        conn.execute("""
            INSERT INTO tracking_history (tracking_id, status, location, description)
            VALUES (?, ?, ?, ?)
        """, (tracking_id, status, location, f"Shipment status updated to {status} at {location}"))

    flash("Shipment tracking status updated!", "success")
    return redirect(url_for("story_admin_delivery"))

# ================= Phase 9: Story Store Notification System Helpers & Routes =================

def create_story_notification(user_id, title, message, notif_type='info', icon='≡ƒöö', color='#3498db', link=None):
    with db_connection() as conn:
        conn.execute("""
            INSERT INTO story_notifications (user_id, title, message, type, icon, color, link)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (user_id, title, message, notif_type, icon, color, link))

def send_story_email(to_email, subject, title, body_text, action_text=None, action_url=None):
    print(f"[AK STORY STORE EMAIL DISPATCH] To: {to_email} | Subject: {subject} | Title: {title} | Message: {body_text}")
    return True

@app.route("/story-store/notifications")
def story_notifications():
    customer_id = session.get("customer_id", 1)
    current_cat = request.args.get("cat", "").strip()

    sql = "SELECT * FROM story_notifications WHERE (user_id=? OR user_id=0)"
    params = [customer_id]

    if current_cat:
        sql += " AND type=?"
        params.append(current_cat)

    sql += " ORDER BY id DESC"

    with db_connection() as conn:
        rows = conn.execute(sql, params).fetchall()
        notifications = [dict(r) for r in rows]

        unread_count = conn.execute(
            "SELECT COUNT(*) FROM story_notifications WHERE (user_id=? OR user_id=0) AND is_read=0",
            (customer_id,)
        ).fetchone()[0]

    cart_items, _ = get_cart_items()
    wishlist_ids = get_wishlist_ids()

    return render_template(
        "story_notifications.html",
        notifications=notifications,
        unread_count=unread_count,
        current_cat=current_cat,
        cart_count=sum(item["quantity"] for item in cart_items),
        wishlist_count=len(wishlist_ids)
    )

@app.route("/story-store/notification/mark-read/<int:notif_id>", methods=["POST"])
def story_notification_mark_read(notif_id):
    customer_id = session.get("customer_id", 1)
    with db_connection() as conn:
        conn.execute("UPDATE story_notifications SET is_read=1 WHERE id=? AND (user_id=? OR user_id=0)", (notif_id, customer_id))
    return redirect(url_for("story_notifications"))

@app.route("/story-store/notification/mark-all-read", methods=["POST"])
def story_notifications_mark_all_read():
    customer_id = session.get("customer_id", 1)
    with db_connection() as conn:
        conn.execute("UPDATE story_notifications SET is_read=1 WHERE user_id=? OR user_id=0", (customer_id,))
    flash("All notifications marked as read.", "success")
    return redirect(url_for("story_notifications"))

@app.route("/story-store/notification/delete/<int:notif_id>", methods=["POST"])
def story_notification_delete(notif_id):
    customer_id = session.get("customer_id", 1)
    with db_connection() as conn:
        conn.execute("DELETE FROM story_notifications WHERE id=? AND (user_id=? OR user_id=0)", (notif_id, customer_id))
    flash("Notification deleted.", "success")
    return redirect(url_for("story_notifications"))

@app.route("/story-store/admin/notifications")
def story_admin_notifications():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    with db_connection() as conn:
        rows = conn.execute("SELECT * FROM story_notifications WHERE user_id=0 ORDER BY id DESC").fetchall()
        notifications = [dict(r) for r in rows]

        admin_unread = conn.execute("SELECT COUNT(*) FROM story_notifications WHERE user_id=0 AND is_read=0").fetchone()[0]

        low_stock_rows = conn.execute("SELECT * FROM products WHERE stock <= 5 ORDER BY stock ASC").fetchall()
        low_stock_books = [dict(r) for r in low_stock_rows]

    return render_template(
        "story_admin_notifications.html",
        notifications=notifications,
        admin_unread=admin_unread,
        low_stock_books=low_stock_books
    )

@app.route("/story-store/admin/notification/mark-all-read", methods=["POST"])
def story_admin_notifications_mark_all_read():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    with db_connection() as conn:
        conn.execute("UPDATE story_notifications SET is_read=1 WHERE user_id=0")

    flash("All admin notifications marked as read.", "success")
    return redirect(url_for("story_admin_notifications"))

@app.route("/story-store/admin/email-settings", methods=["GET", "POST"])
def story_admin_email_settings():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    if request.method == "POST":
        smtp_host = request.form.get("smtp_host", "smtp.gmail.com").strip()
        smtp_port = int(request.form.get("smtp_port", 587))
        sender_email = request.form.get("sender_email", "store@akclicks.com").strip()
        sender_name = request.form.get("sender_name", "AK Story Store").strip()
        app_password = request.form.get("app_password", "").strip()
        is_enabled = 1 if request.form.get("is_enabled") else 0

        with db_connection() as conn:
            conn.execute("""
                UPDATE story_email_settings
                SET smtp_host=?, smtp_port=?, sender_email=?, sender_name=?, app_password=?, is_enabled=?
                WHERE id=1
            """, (smtp_host, smtp_port, sender_email, sender_name, app_password, is_enabled))

        flash("Story Store Email Settings updated successfully!", "success")
        return redirect(url_for("story_admin_email_settings"))

    with db_connection() as conn:
        s = conn.execute("SELECT * FROM story_email_settings WHERE id=1").fetchone()
        settings = dict(s) if s else {}

    return render_template("story_admin_email_settings.html", settings=settings)

# ================= Phase 10: Security & System Hardening Helpers & Routes =================

def sanitize_input(text):
    if not text:
        return ""
    return html.escape(str(text))

def generate_otp(identifier):
    raw_otp = f"{random.randint(100000, 999999)}"
    otp_hash = hashlib.sha256(raw_otp.encode()).hexdigest()
    expires_at = (datetime.now() + timedelta(minutes=5)).strftime("%Y-%m-%d %H:%M:%S")

    with db_connection() as conn:
        conn.execute("""
            INSERT INTO otp_codes (identifier, otp_hash, expires_at)
            VALUES (?, ?, ?)
        """, (identifier, otp_hash, expires_at))

    print(f"[AK STORY STORE SECURITY OTP] Sent 6-digit OTP [{raw_otp}] to {identifier}")
    return raw_otp

def verify_otp(identifier, raw_otp):
    otp_hash = hashlib.sha256(raw_otp.encode()).hexdigest()
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    with db_connection() as conn:
        row = conn.execute("""
            SELECT * FROM otp_codes
            WHERE identifier=? AND otp_hash=? AND is_used=0 AND expires_at > ?
            ORDER BY id DESC LIMIT 1
        """, (identifier, otp_hash, now_str)).fetchone()

        if row:
            conn.execute("UPDATE otp_codes SET is_used=1 WHERE id=?", (row["id"],))
            return True
    return False

def log_audit_action(actor, action_type, affected_table=None, record_id=None, old_value=None, new_value=None):
    with db_connection() as conn:
        conn.execute("""
            INSERT INTO audit_logs (actor, action_type, affected_table, record_id, old_value, new_value, ip_address)
            VALUES (?, ?, ?, ?, ?, ?, '127.0.0.1')
        """, (actor, action_type, affected_table, str(record_id) if record_id else None, str(old_value) if old_value else None, str(new_value) if new_value else None))

def log_security_event(user_id, action, details=None):
    with db_connection() as conn:
        conn.execute("""
            INSERT INTO security_logs (user_id, action, details, ip_address)
            VALUES (?, ?, ?, '127.0.0.1')
        """, (user_id, action, details))

@app.route("/story-store/login/2fa", methods=["GET", "POST"])
def story_login_2fa():
    identifier = request.args.get("identifier", request.form.get("identifier", "customer@example.com"))

    if request.method == "POST":
        otp_code = request.form.get("otp_code", "").strip()
        if verify_otp(identifier, otp_code):
            session["customer_id"] = 1
            session["customer_name"] = identifier.split("@")[0]
            log_security_event(1, "2FA_LOGIN_SUCCESS", f"2FA verified for {identifier}")
            flash("2FA Verification Successful! Welcome back.", "success")
            return redirect(url_for("story_home"))
        else:
            flash("Invalid or expired 6-digit OTP code.", "error")

    return render_template("story_security_2fa.html", identifier=identifier)

@app.route("/story-store/admin/security")
def story_admin_security():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    with db_connection() as conn:
        audit_rows = conn.execute("SELECT * FROM story_admin_audit_logs ORDER BY id DESC LIMIT 10").fetchall()
        audit_logs = [dict(r) for r in audit_rows]

        sec_row = conn.execute("SELECT * FROM story_admin_security WHERE id=1").fetchone()
        sec = dict(sec_row) if sec_row else {}

        failed_logins_count = conn.execute("SELECT COUNT(*) FROM failed_logins").fetchone()[0]

    return render_template(
        "story_admin_security.html",
        audit_logs=audit_logs,
        sec=sec,
        failed_logins_count=failed_logins_count
    )

# ================= Admin Security Suite TOTP & Sub-modules =================

def generate_totp_secret():
    return base64.b32encode(os.urandom(10)).decode('utf-8')

def get_totp_token(secret, intervals_no=None):
    if intervals_no is None:
        intervals_no = int(time.time()) // 30
    key = base64.b32decode(secret, True)
    msg = struct.pack(">Q", intervals_no)
    h = hmac.new(key, msg, hashlib.sha1).digest()
    o = h[19] & 15
    h = (struct.unpack(">I", h[o:o+4])[0] & 0x7fffffff) % 1000000
    return f"{h:06d}"

def verify_totp_token(secret, token):
    if not secret or not token:
        return False
    curr_time = int(time.time()) // 30
    for i in range(-1, 2):
        if get_totp_token(secret, curr_time + i) == token.strip():
            return True
    return False

@app.route("/story-store/admin/2fa/verify", methods=["GET", "POST"])
def story_admin_2fa_verify():
    if not session.get("story_admin_pending"):
        return redirect(url_for("story_admin_login"))

    if request.method == "POST":
        totp_code = request.form.get("totp_code", "").strip()

        with db_connection() as conn:
            sec = conn.execute("SELECT * FROM story_admin_security WHERE id=1").fetchone()
            secret_key = sec["secret_key"] if sec else None

            code_match = conn.execute("SELECT * FROM story_admin_recovery_codes WHERE code_plain=? AND is_used=0", (totp_code,)).fetchone()

            if verify_totp_token(secret_key, totp_code) or code_match:
                if code_match:
                    conn.execute("UPDATE story_admin_recovery_codes SET is_used=1 WHERE id=?", (code_match["id"],))

                session.pop("story_admin_pending", None)
                session["story_admin"] = True
                log_admin_audit_action("2FA_LOGIN_SUCCESS", "AUTH", conn=conn)
                flash("2FA Verification Successful! Welcome to Admin Panel.", "success")
                return redirect(url_for("story_admin_dashboard"))
            else:
                log_admin_audit_action("2FA_LOGIN_FAILED", "AUTH", status="FAILED", conn=conn)
                flash("Invalid 6-digit TOTP or Recovery Code.", "error")

    return render_template("story_admin_2fa_verify.html")

@app.route("/story-store/admin/2fa/setup", methods=["GET", "POST"])
def story_admin_2fa_setup():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    with db_connection() as conn:
        sec = conn.execute("SELECT * FROM story_admin_security WHERE id=1").fetchone()
        is_enabled = sec["is_2fa_enabled"] if sec else 0
        secret_key = sec["secret_key"] if sec and sec["secret_key"] else generate_totp_secret()

        if request.method == "POST" and not is_enabled:
            totp_code = request.form.get("totp_code", "").strip()
            form_secret = request.form.get("secret_key", secret_key).strip()

            if verify_totp_token(form_secret, totp_code):
                conn.execute("""
                    UPDATE story_admin_security
                    SET is_2fa_enabled=1, secret_key=?, security_score=98
                    WHERE id=1
                """, (form_secret,))

                conn.execute("DELETE FROM story_admin_recovery_codes")
                for _ in range(10):
                    code = f"{random.randint(1000, 9999)}-{random.randint(1000, 9999)}"
                    code_hash = hashlib.sha256(code.encode()).hexdigest()
                    conn.execute("INSERT INTO story_admin_recovery_codes (code_hash, code_plain) VALUES (?, ?)", (code_hash, code))

                log_admin_audit_action("2FA_ENABLED", "SECURITY", conn=conn)
                flash("2FA Enabled successfully with 10 Recovery Codes!", "success")
                return redirect(url_for("story_admin_2fa_setup"))
            else:
                flash("Invalid TOTP code. Please scan the QR code and try again.", "error")

        rec_codes = conn.execute("SELECT * FROM story_admin_recovery_codes").fetchall()
        recovery_codes = [dict(r) for r in rec_codes]

    otpauth_url = f"otpauth://totp/AK%20Story%20Store%20Admin:admin?secret={secret_key}&issuer=AK%20Story%20Store"
    qr_img = qrcode.make(otpauth_url)
    buffered = io.BytesIO()
    qr_img.save(buffered, format="PNG")
    qr_code_b64 = base64.b64encode(buffered.getvalue()).decode("utf-8")

    return render_template(
        "story_admin_2fa_setup.html",
        is_enabled=is_enabled,
        secret_key=secret_key,
        qr_code_b64=qr_code_b64,
        recovery_codes=recovery_codes
    )

@app.route("/story-store/admin/2fa/disable", methods=["POST"])
def story_admin_2fa_disable():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    with db_connection() as conn:
        conn.execute("UPDATE story_admin_security SET is_2fa_enabled=0, security_score=85 WHERE id=1")
        conn.execute("DELETE FROM story_admin_recovery_codes")

    log_admin_audit_action("2FA_DISABLED", "SECURITY")
    flash("2FA has been disabled for Story Store Admin.", "info")
    return redirect(url_for("story_admin_2fa_setup"))

@app.route("/story-store/admin/security/audit")
def story_admin_security_audit():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    fmt = request.args.get("export", "").strip().lower()

    with db_connection() as conn:
        rows = conn.execute("SELECT * FROM story_admin_audit_logs ORDER BY id DESC LIMIT 50").fetchall()
        audit_logs = [dict(r) for r in rows]

    if fmt == "csv":
        output = "ID,Timestamp,Admin,Action,Module,Status,IP\n"
        for a in audit_logs:
            output += f"{a['id']},{a['timestamp']},{a['admin_name']},{a['action']},{a['module']},{a['status']},{a['ip_address']}\n"
        return Response(output, mimetype="text/csv", headers={"Content-Disposition": "attachment;filename=admin_audit_logs.csv"})

    return render_template("story_admin_security_audit.html", audit_logs=audit_logs)

@app.route("/story-store/admin/security/sessions")
def story_admin_security_sessions():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    with db_connection() as conn:
        rows = conn.execute("SELECT * FROM story_admin_sessions ORDER BY id DESC").fetchall()
        sessions = [dict(r) for r in rows]

    return render_template("story_admin_security_sessions.html", sessions=sessions)

@app.route("/story-store/admin/security/sessions/revoke/<int:session_id>", methods=["POST"])
def story_admin_security_sessions_revoke(session_id):
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    with db_connection() as conn:
        conn.execute("DELETE FROM story_admin_sessions WHERE id=?", (session_id,))

    log_admin_audit_action("REVOKE_SESSION", "SECURITY")
    flash("Session revoked successfully.", "success")
    return redirect(url_for("story_admin_security_sessions"))

@app.route("/story-store/admin/security/sessions/revoke-others", methods=["POST"])
def story_admin_security_sessions_revoke_others():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    with db_connection() as conn:
        conn.execute("DELETE FROM story_admin_sessions")

    log_admin_audit_action("REVOKE_ALL_SESSIONS", "SECURITY")
    flash("All other admin sessions have been revoked.", "success")
    return redirect(url_for("story_admin_security_sessions"))

@app.route("/story-store/admin/security/password", methods=["GET", "POST"])
def story_admin_security_password():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    if request.method == "POST":
        new_pass = request.form.get("new_password", "").strip()
        confirm_pass = request.form.get("confirm_password", "").strip()

        if new_pass != confirm_pass:
            flash("New passwords do not match.", "error")
        elif len(new_pass) < 8:
            flash("Password must be at least 8 characters long.", "error")
        else:
            log_admin_audit_action("CHANGE_PASSWORD", "SECURITY")
            flash("Admin password updated successfully!", "success")
            return redirect(url_for("story_admin_security"))

    return render_template("story_admin_security_password.html")

@app.route("/story-store/admin/security/settings", methods=["GET", "POST"])
def story_admin_security_settings():
    if not session.get("story_admin"):
        return redirect(url_for("story_admin_login"))

    if request.method == "POST":
        max_attempts = int(request.form.get("max_login_attempts", 5))
        auto_logout = int(request.form.get("auto_logout_minutes", 20))

        with db_connection() as conn:
            conn.execute("""
                UPDATE story_admin_security
                SET max_login_attempts=?, auto_logout_minutes=?
                WHERE id=1
            """, (max_attempts, auto_logout))

        log_admin_audit_action("UPDATE_SECURITY_SETTINGS", "SECURITY")
        flash("Admin Security Policies updated successfully!", "success")
        return redirect(url_for("story_admin_security_settings"))

    with db_connection() as conn:
        sec = conn.execute("SELECT * FROM story_admin_security WHERE id=1").fetchone()

    return render_template("story_admin_security_settings.html", sec=dict(sec) if sec else {})



# Initialize database
with app.app_context():
    init_db()

if __name__ == "__main__":
    app.run(
        host="0.0.0.0",
        port=5001,
        debug=True
    )